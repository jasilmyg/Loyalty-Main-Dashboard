import os
import threading
# psycopg2 is optional for local development (may not be installed)
try:
    import psycopg2
except Exception:
    psycopg2 = None
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from .services import AnalyticsService

# ── Singleton AnalyticsService ─────────────────────────────────────────────────
# Creating a new DuckDB+PostgreSQL connection on every request costs 2-5 seconds.
# We create ONE instance at startup and reuse it safely across threads.
_svc_instance = None
_svc_lock = threading.Lock()

def get_analytics():
    """Return the shared AnalyticsService instance (created once at startup)."""
    global _svc_instance
    if _svc_instance is None:
        with _svc_lock:
            if _svc_instance is None:
                _svc_instance = AnalyticsService()
    return _svc_instance

def get_filters(request):
    """Extract standard filters from GET parameters."""
    # Support multi-branch: ?branches=BranchA,BranchB or repeated ?branch=...
    branches_param = request.GET.get('branches', '')
    branch_param   = request.GET.get('branch', '')

    filters = {
        'start_date': request.GET.get('start_date'),
        'end_date':   request.GET.get('end_date'),
        'branch':     branch_param,
        'branches':   branches_param,   # comma-separated multi-branch
        'staff':      request.GET.get('staff'),
        'rbm':        request.GET.get('rbm'),
        'bdm':        request.GET.get('bdm'),
        'period':     request.GET.get('period', 'monthly'),
    }

    # Simple Role-Based Access Control logic
    user = request.user
    if user.is_authenticated:
        if user.role == 'Staff' and user.branch:
            filters['branch'] = user.branch
            filters['branches'] = user.branch
    return filters

class ClickHouseHealthAPI(APIView):
    """Diagnostic endpoint — /api/v1/ch-health/ — no auth required."""
    permission_classes = []

    def get(self, request):
        import time, os
        from analytics.clickhouse_service import CH_HOST, CH_PORT, CH_USER, CH_DATABASE

        result = {
            'ch_host':     CH_HOST,
            'ch_port':     CH_PORT,
            'ch_user':     CH_USER,
            'ch_database': CH_DATABASE,
            'ch_password_set': bool(os.environ.get('CH_PASSWORD')),
        }

        t0 = time.time()
        try:
            try:
                import clickhouse_connect
                test_client = clickhouse_connect.get_client(
                    host=CH_HOST,
                    port=CH_PORT,
                    username=CH_USER,
                    password=os.environ.get('CH_PASSWORD', 'ZFlujj9SA_Iei'),
                    database=CH_DATABASE,
                    secure=True,
                    connect_timeout=5,
                )
                rows = test_client.query('SELECT COUNT(*) FROM sales_data').result_rows
                result['status']     = 'OK'
                result['row_count']  = rows[0][0] if rows else 0
                result['elapsed_ms'] = round((time.time() - t0) * 1000)
            except Exception as inner_e:
                result['status'] = 'FAILED'
                result['error'] = str(inner_e)
        except Exception as e:
            result['status']     = 'ERROR'
            result['error']      = str(e)
            result['elapsed_ms'] = round((time.time() - t0) * 1000)

        return Response(result)


class SalesOverviewAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_sales_overview(get_filters(request))
        return Response(data)


class CategoryAnalysisAPI(APIView):
    """
    Returns category-wise revenue and quantity from Azure Blob data via ClickHouse.
    """
    permission_classes = [IsAuthenticated]
    def get(self, request):
        filters = get_filters(request)
        data = get_analytics().get_category_analysis(filters)
        return Response({'success': True, 'data': data})

class CustomerAnalyticsAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_customer_analytics(get_filters(request))
        return Response(data)

class RFMAnalysisAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().perform_rfm_analysis(get_filters(request))
        return Response(data)

class MonetaryQuintilesAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_monetary_quintiles(get_filters(request))
        return Response(data)

class CohortRetentionAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_cohort_retention()
        return Response(data)

class YearlyCohortAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_yearly_cohort_analysis()
        return Response(data)

class PaymentAnalyticsAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_payment_analytics(get_filters(request))
        return Response(data)

class DiscountAnalysisAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_discount_analysis(get_filters(request))
        return Response(data)

class StaffPerformanceAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_staff_performance(get_filters(request))
        return Response(data)

class BranchPerformanceAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_branch_performance(get_filters(request))
        return Response(data)

class FrequencyDistributionAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_frequency_distribution(get_filters(request))
        return Response(data)

class LoyaltyOverviewAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_loyalty_overview_kpis(get_filters(request))
        return Response(data)

class GapAnalysisAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_gap_segmentation(get_filters(request))
        return Response(data)

class LoyaltySegmentationAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_customer_segmentation_matrix(get_filters(request))
        return Response(data)

class ActionEngineAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_action_engine_data(get_filters(request))
        return Response(data)


class BusinessInsightsAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        svc = get_analytics()
        insight_type = request.GET.get('type')
        if insight_type == 'cohort':
            data = svc.get_cohort_business_insights()
        else:
            data = svc.get_business_insights(get_filters(request))
        return Response(data)

class RetailLoyaltyReportAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_retail_loyalty_report(get_filters(request))
        return Response(data)

class RetailLoyaltyAdvancedReportAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        return Response(get_analytics().get_retail_loyalty_advanced_report(get_filters(request)))

class FYLoyaltyReportAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_fy_loyalty_report(get_filters(request))
        return Response(data)

class FYSalesReportAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_fy_sales_report(get_filters(request))
        return Response(data)

class InvalidMobilesAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_invalid_mobiles()
        return Response(data)

class BranchesAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data = get_analytics().get_unique_branches()
        return Response(data)

try:
    import psycopg2
except Exception:
    psycopg2 = None

class DBManagerAPI(APIView):
    """
    Paginated full-data viewer for the DB Manager section.
    Fetches data DIRECTLY from remote PostgreSQL via psycopg2.
    Returns:  { columns, rows, page, page_size, total_rows, total_pages }
    """
    permission_classes = [IsAuthenticated]
    DB_PASSWORD = os.environ.get('DB_MANAGER_PASSWORD', 'myGLoyalty@2024')
    PAGE_SIZE   = 100

    def get_pg_conn(self):
        import django.conf
        db = django.conf.settings.DATABASES['default']
        return psycopg2.connect(
            host=db['HOST'],
            port=db['PORT'],
            dbname=db['NAME'],
            user=db['USER'],
            password=db['PASSWORD'],
            sslmode='require',
            connect_timeout=10
        )

    def get(self, request):
        # ── Password gate ──────────────────────────────────────────────
        pwd = request.GET.get('db_password', '')
        if pwd != self.DB_PASSWORD:
            return Response({'error': 'Unauthorized', 'detail': 'Invalid password'}, status=403)

        # ── Pagination ─────────────────────────────────────────────────
        try:
            page = max(1, int(request.GET.get('page', 1)))
        except (ValueError, TypeError):
            page = 1

        page_size = self.PAGE_SIZE
        offset    = (page - 1) * page_size

        # ── Search / filter ────────────────────────────────────────────
        search     = (request.GET.get('search') or '').strip()
        col_filter = (request.GET.get('col')    or '').strip()

        where_clause = "1=1"
        params = []
        if search:
            like_val = f'%{search}%'
            if col_filter:
                where_clause = f'CAST("{col_filter}" AS VARCHAR) ILIKE %s'
                params = [like_val]
            else:
                where_clause = (
                    '"Customer Name"    ILIKE %s OR '
                    'CAST("Customer Mobile" AS VARCHAR) ILIKE %s OR '
                    '"Invoice Number"   ILIKE %s OR '
                    '"Branch"           ILIKE %s'
                )
                params = [like_val] * 4

        try:
            conn = self.get_pg_conn()
            cur  = conn.cursor()

            # ── Total count ────────────────────────────────────────────
            count_sql = f'SELECT COUNT(*) FROM sales_data WHERE {where_clause}'
            cur.execute(count_sql, params)
            total_rows  = cur.fetchone()[0] or 0
            total_pages = max(1, -(-total_rows // page_size))

            # ── Data fetch ─────────────────────────────────────────────
            data_sql = (
                f'SELECT * FROM sales_data WHERE {where_clause} '
                f'ORDER BY "Date" DESC '
                f'LIMIT {page_size} OFFSET {offset}'
            )
            cur.execute(data_sql, params)
            columns = [desc[0] for desc in cur.description]
            rows    = cur.fetchall()

            cur.close()
            conn.close()

            # Serialize: convert any non-JSON-safe types to strings
            def _safe(v):
                if v is None:
                    return None
                if isinstance(v, (int, float, str, bool)):
                    return v
                return str(v)

            return Response({
                'columns':     columns,
                'rows':        [[_safe(c) for c in row] for row in rows],
                'page':        page,
                'page_size':   page_size,
                'total_rows':  total_rows,
                'total_pages': total_pages,
            })

        except Exception as exc:
            import traceback
            return Response(
                {'error': 'DB query failed', 'detail': str(exc)},
                status=500
            )


import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _build_xlsx_response(filename, headers, rows):
    """Build a styled .xlsx HttpResponse — plain Django, no DRF wrapping."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()

    response = HttpResponse(
        content=xlsx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # RFC 5987 encoding keeps Unicode filenames safe across browsers
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = str(len(xlsx_bytes))
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response


def _apply_multi_branch_to_filters(filters):
    """
    If 'branches' param contains multiple branches (comma-separated),
    temporarily set 'branch' to a sentinel and return a branch list.
    Returns (updated_filters, branch_list_or_None).
    """
    branches_raw = filters.get('branches', '')
    if branches_raw:
        branch_list = [b.strip() for b in branches_raw.split(',') if b.strip()]
        if len(branch_list) == 1:
            # Single branch — use normal single-branch path
            f = dict(filters)
            f['branch'] = branch_list[0]
            f['branches'] = ''
            return f, None
        elif len(branch_list) > 1:
            # Multi-branch: clear single branch so _build_where_clause doesn't add it
            f = dict(filters)
            f['branch'] = ''
            f['branches'] = ''
            return f, branch_list
    return filters, None


def export_view(request, module):
    """Plain Django download view — avoids DRF response wrapping that can corrupt binary files."""
    import json, math
    svc = get_analytics()
    filters = get_filters(request)

    # ── Resolve multi-branch for this export ──────────────────────────────────
    filters, multi_branches = _apply_multi_branch_to_filters(filters)
    where_sql, params = svc._build_where_clause(filters)

    # If multiple branches selected, inject an IN clause
    if multi_branches:
        placeholders = ', '.join(['%s'] * len(multi_branches))
        upper_list   = [b.upper() for b in multi_branches]
        branch_cond  = f'UPPER("Branch") IN ({placeholders})'
        if where_sql == '1=1':
            where_sql = branch_cond
        else:
            where_sql = f'{where_sql} AND {branch_cond}'
        params = list(params) + upper_list

    if module == 'customer-frequency':
        rows_data = svc.get_frequency_distribution(filters)
        headers = ['Segment', 'Customers', 'Net Revenue (INR)', 'Customer %', 'Revenue %', 'ASP (INR)']
        rows = [
            (r['segment'], r['customers'], r['net_revenue'],
             r['cust_pct'], r['rev_pct'], r['asp'])
            for r in rows_data
        ]
        return _build_xlsx_response('customer_frequency_report.xlsx', headers, rows)

    elif module in ('rfm', 'rfm-segments'):
        segment = request.GET.get('segment')
        if not segment:
            # Summary Export
            rows_data = svc.get_rfm_segments(filters)
            headers = ['RFM Segment', 'Customer Count', 'Total Revenue (INR)', 'Average Revenue (INR)']
            rows = [
                (r['segment'], r['count'], r['total_revenue'], r['avg_revenue'])
                for r in rows_data
            ]
            return _build_xlsx_response('rfm_summary_report.xlsx', headers, rows)
        else:
            # Detail Export for a specific segment — respect multi-branch filter
            from django.db import connection as db_conn
            query, q_params = svc.get_rfm_details_query(filters, segment)

            # If multi-branch was selected, rebuild query with branch IN clause
            if multi_branches:
                query, q_params = _build_rfm_segment_query_multi_branch(
                    svc, filters, multi_branches, segment
                )

            query += " LIMIT 100000"
            with db_conn.cursor() as cur:
                cur.execute(query, q_params)
                headers = [col[0] for col in cur.description]
                rows = cur.fetchall()

            filename = f"rfm_{segment.lower().replace(' ', '_')}_details.xlsx"
            return _build_xlsx_response(filename, headers, rows)

    elif module == 'sales':
        from django.db import connection as db_conn
        table = "sales_data"
        query = f'SELECT * FROM {table} WHERE {where_sql} LIMIT 50000'
        with db_conn.cursor() as cur:
            cur.execute(query, params)
            headers = [col[0] for col in cur.description]
            rows = cur.fetchall()
        return _build_xlsx_response('sales_export.xlsx', headers, rows)

    elif module == 'gap-analysis':
        rows_data = svc.get_gap_segmentation(filters)
        headers = ['Gap Range', 'Customers', 'Percentage (%)', 'Avg Gap (Days)', 'Loyalty Signal', 'Priority', 'Action Strategy']
        rows = [
            (r['segment'], r['count'], r['percentage'], r['avg_gap'], r['signal'], r['priority'], r.get('action', ''))
            for r in rows_data
        ]
        return _build_xlsx_response('gap_analysis_report.xlsx', headers, rows)

    elif module == 'segment-customers':
        segment = request.GET.get('segment', '')
        part    = int(request.GET.get('part', '1'))

        if not segment:
            return HttpResponse("segment parameter required", status=400)

        chunk = svc.SEGMENT_CHUNK_SIZE

        # part=0  → return JSON describing how many parts exist
        if part == 0:
            total = svc.count_customers_for_segment(filters, segment)
            total_parts = max(1, math.ceil(total / chunk))
            return HttpResponse(
                json.dumps({"total_parts": total_parts, "total_customers": total}),
                content_type='application/json'
            )

        # part=N → return the actual Excel file for that chunk
        offset  = (part - 1) * chunk
        headers, rows = svc.get_customers_for_segment(filters, segment, offset)

        # Make the filename filesystem-safe
        safe_seg = segment.replace(' ', '-').replace('/', '-')
        filename = f"customers_{safe_seg}_part{part}.xlsx"

        # Friendly column headers for Excel
        display_headers = ['Customer Mobile', 'Customer Name', 'Visits',
                           'Net Revenue (INR)', 'Last Visit Date']
        return _build_xlsx_response(filename, display_headers, rows)

    elif module == 'all-customers':
        part = int(request.GET.get('part', '1'))
        chunk = svc.SEGMENT_CHUNK_SIZE

        if part == 0:
            total = svc.count_all_customers(filters)
            total_parts = max(1, math.ceil(total / chunk))
            return HttpResponse(
                json.dumps({"total_parts": total_parts, "total_customers": total}),
                content_type='application/json'
            )

        offset = (part - 1) * chunk
        headers, rows = svc.get_all_customers(filters, offset)

        filename = f"all_customers_part{part}.xlsx" if svc.count_all_customers(filters) > chunk else "all_customers.xlsx"
        display_headers = ['Customer Mobile', 'Customer Name', 'Visits', 'Net Revenue (INR)', 'Last Visit Date']
        return _build_xlsx_response(filename, display_headers, rows)

    elif module == 'retail-analytics':
        rows_data, db_start = svc.get_retail_loyalty_matrix(filters)

        period = filters.get('period', 'monthly')
        col_period = "Year" if period == "yearly" else ("Quarter" if period == "quarterly" else "Month")
        col_mom = "YoY" if period == "yearly" else ("QoQ" if period == "quarterly" else "MoM")

        headers = [col_period, 'Total Members', f'{col_mom} Members %',
                   'New Members', 'Repeat Members', 'Repeat %', 'Retention % (DB)', 'Cumulative DB']

        import calendar
        rows = []
        rolling_db = db_start
        prev_final_db = rolling_db

        for r in rows_data:
            raw_period = r.get('month', '')
            display_period = raw_period

            if period == 'monthly' and '-' in raw_period:
                try:
                    y, m = raw_period.split('-')
                    display_period = f"{calendar.month_name[int(m)]} {y}"
                except Exception:
                    pass
            elif period == 'quarterly' and '-Q' in raw_period:
                try:
                    y, q = raw_period.split('-Q')
                    q_map = {'1': 'JFM', '2': 'AMJ', '3': 'JAS', '4': 'OND'}
                    display_period = f"{q_map.get(q, 'Q'+q)} {y}"
                except Exception:
                    pass

            rolling_db += r.get('new_members', 0)
            final_db = r.get('db_size', 0) if r.get('db_size', 0) > 0 else rolling_db

            retention_pct = (r.get('repeat_members', 0) / prev_final_db * 100) if prev_final_db > 0 else 0
            prev_final_db = final_db

            mom_str = f"{r.get('mom_total_members', 0)}%" if r.get('mom_total_members', 0) else "—"

            rows.append((
                display_period,
                r.get('total_members', 0),
                mom_str,
                r.get('new_members', 0),
                r.get('repeat_members', 0),
                f"{r.get('repeat_pct', 0)}%",
                f"{round(retention_pct, 2)}%",
                final_db
            ))

        return _build_xlsx_response(f'retail_loyalty_analytics_{period}_report.xlsx', headers, rows)

    elif module == 'invalid-mobiles':
        data = svc.get_invalid_mobiles()
        headers = ['Raw Mobile', 'Customer Name', 'Branch', 'Sale Date', 'Invoice Number']
        rows = [
            (r['raw_mobile'], r['customer_name'], r['branch'], r['sale_date'], r['invoice_number'])
            for r in data['rows']
        ]
        return _build_xlsx_response('invalid_mobiles.xlsx', headers, rows)

    elif module == 'fy-loyalty-report':
        data = svc.get_fy_loyalty_report(filters)
        headers = ['Financial Year', 'Total Members', 'YoY Members %', 'New Members', 'Repeat Members', 'Repeat %', 'Retention % (DB)', 'Cumulative DB']
        rows = [
            (
                r['fy_label'],
                r['total_members'],
                r['yoy_pct'],
                r['new_members'],
                r['repeat_members'],
                r['repeat_pct'],
                r['retention_pct_db'],
                r['cumulative_db']
            )
            for r in data
        ]
        return _build_xlsx_response('fy_loyalty_report.xlsx', headers, rows)

    elif module == 'fy-sales-report':
        data = svc.get_fy_sales_report(filters)
        headers = ['Financial Year', 'Total Sale (Cr)', 'YoY Sale Growth %', 'New Members Sale (Cr)', 'Repeat Members Sale (Cr)', 'Repeat Sale %', 'ASP (Customer)']
        rows = [
            (
                r['fy_label'],
                r['total_sale_cr'],
                r['yoy_sale_pct'],
                r['new_member_sale_cr'],
                r['repeat_member_sale_cr'],
                r['repeat_sale_pct'],
                r['asp']
            )
            for r in data
        ]
        return _build_xlsx_response('fy_sales_report.xlsx', headers, rows)

    return HttpResponse("Unknown export module", status=400)


def _build_rfm_segment_query_multi_branch(svc, filters, multi_branches, segment):
    """
    Build the RFM details query with a multi-branch IN filter injected
    into the CTE base scan instead of a single-branch equality.
    """
    from .services import TABLE, VALID_MOBILE

    placeholders = ', '.join(['%s'] * len(multi_branches))
    upper_list   = [b.upper() for b in multi_branches]
    branch_cond  = f'UPPER("Branch") IN ({placeholders})'

    # Build extra WHERE conditions (date range, staff, rbm, bdm) — without branch
    branch_less_filters = dict(filters)
    branch_less_filters['branch'] = ''
    where_sql, base_params = svc._build_where_clause(branch_less_filters)

    # Inject branch IN clause
    if where_sql == '1=1':
        combined_where = branch_cond
    else:
        combined_where = f'{where_sql} AND {branch_cond}'

    combined_params = list(base_params) + upper_list

    # Use the slow-path CTE (forces raw v_sales_data scan so Branch is available)
    cte = f"""
        WITH rfm_base AS (
            SELECT "Customer Mobile" AS mobile,
                MAX("Customer Name")              AS customer_name,
                (CURRENT_DATE - MAX("Date"))::INT AS recency,
                COUNT(DISTINCT "Date")            AS frequency,
                SUM("Total Value")::FLOAT         AS monetary,
                MAX("Date")                       AS last_visit
            FROM {TABLE}
            WHERE {combined_where}
              AND "Customer Mobile" ~ '^[0-9]{{10}}$'
            GROUP BY "Customer Mobile"
        ),
        scored AS (
            SELECT *,
                CASE WHEN recency<=90 THEN 5 WHEN recency<=180 THEN 4
                     WHEN recency<=365 THEN 3 WHEN recency<=730 THEN 2 ELSE 1 END AS r_score,
                CASE WHEN frequency>=5 THEN 5 WHEN frequency=4 THEN 4
                     WHEN frequency=3 THEN 3 WHEN frequency=2 THEN 2 ELSE 1 END AS f_score,
                NTILE(5) OVER (ORDER BY monetary ASC) AS m_score
            FROM rfm_base
        ),
        segmented AS (
            SELECT *,
                r_score::TEXT||f_score::TEXT||m_score::TEXT AS rfm_code,
                CASE
                    WHEN r_score>=4 AND f_score>=4 AND m_score>=4 THEN 'Champions'
                    WHEN r_score>=3 AND f_score>=3 AND m_score>=3 THEN 'Loyal'
                    WHEN r_score>=4 AND f_score<=2               THEN 'New'
                    WHEN r_score=2 AND f_score>=3 AND m_score>=3 THEN 'At Risk'
                    WHEN r_score=1                               THEN 'Lost'
                    ELSE 'Others'
                END AS segment
            FROM scored
        )
    """

    query = f"""
        {cte}
        SELECT customer_name AS "Customer Name", mobile AS "Customer Mobile",
            recency AS "Recency (Days)", frequency AS "Frequency (Visits)",
            monetary AS "Monetary Value", r_score AS "R Score",
            f_score AS "F Score", m_score AS "M Score",
            rfm_code AS "RFM Code", segment AS "RFM Segment",
            last_visit AS "Last Visit Date"
        FROM segmented WHERE segment = %s
        ORDER BY monetary DESC NULLS LAST
    """
    return query, combined_params + [segment]


# Keep the old DRF class so existing URL patterns still resolve without errors
class ExportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, module):
        return export_view(request, module)
