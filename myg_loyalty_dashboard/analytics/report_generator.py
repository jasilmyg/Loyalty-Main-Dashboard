import io
import zipfile
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

CRORE = 10_000_000

# ── CATEGORY MAPPING ─────────────────────────────────────────────────────────
CAT_MAP = {
    ("TELECOM",              "MOBILE")  : "MOBILE",
    ("TELECOM",              "TABLET")  : "TABLET",
    ("IT",                   "LAPTOP")  : "LAPTOP",
    ("IT",                   "RIG")     : "RIG",
    ("IT",                   "DESKTOP") : "DESKTOP",
    ("IT",                   "PRINTER") : "LAPTOP",
    ("ACCESSORIES",          "*")       : "ACC",
    ("CONSUMER ELECTRONICS", "*")       : "CE",
    ("VALUE ADDED SERVICE",  "*")       : "VAS",
    ("ENDPOINT PROTECTION",  "*")       : "OTHER",
    ("HOME SOLUTIONS",       "*")       : "OTHER",
    ("OTHERS",               "*")       : "OTHER",
}
PDT_CAT_ORDER = ["MOBILE", "CE", "LAPTOP", "ACC", "OTHER", "TABLET", "RIG", "DESKTOP"]

# ── STYLE CONSTANTS ───────────────────────────────────────────────────────────
THIN       = Side(style="thin",   color="000000")
THIN_BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HDR_FONT   = Font(name="Arial", bold=True,  size=9)
BODY_FONT  = Font(name="Arial", bold=False, size=9)
TOT_FONT   = Font(name="Arial", bold=True,  size=9)

NEG_FILL   = PatternFill("solid", fgColor="FFC7CE")
TOT_FILL   = PatternFill("solid", fgColor="BDD7EE")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

FMT_AMT = '#,##0.00'
FMT_QTY = '#,##0'
FMT_PCT = '0.00"%"'

from django.db.models import Sum
from analytics.models import ProductSale

def load_from_db(year, month):
    qs = ProductSale.objects.filter(date__year=year, date__month=month).values(
        'product', 'category', 'brand'
    ).annotate(
        total_qty=Sum('qty'),
        total_price=Sum('sold_price')
    )
    df = pd.DataFrame(list(qs))
    if df.empty:
        df = pd.DataFrame(columns=['Product', 'Category', 'Brand', 'QTY', 'Sold Price'])
    else:
        df = df.rename(columns={
            'product': 'Product',
            'category': 'Category',
            'brand': 'Brand',
            'total_qty': 'QTY',
            'total_price': 'Sold Price'
        })
        df['Sold Price'] = pd.to_numeric(df['Sold Price'], errors='coerce').fillna(0)
        df['QTY'] = pd.to_numeric(df['QTY'], errors='coerce').fillna(0)
    
    df["Category"]   = df["Category"].astype(str).str.strip().str.upper()
    df["Product"]    = df["Product"].astype(str).str.strip().str.upper()
    df["Brand"]      = df["Brand"].astype(str).str.strip().str.upper()
    return df

def add_pdt_cat(df):
    def classify(row):
        cat, prd = row["Category"], row["Product"]
        if (cat, prd) in CAT_MAP:
            return CAT_MAP[(cat, prd)]
        if (cat, "*") in CAT_MAP:
            return CAT_MAP[(cat, "*")]
        return "OTHER"
    df = df.copy()
    df["PDT_CAT"] = df.apply(classify, axis=1)
    return df

def pivot(apr, may, group_col):
    def agg(df):
        return df.groupby(group_col, as_index=False).agg(
            AMT=("Sold Price", "sum"),
            QTY=("QTY",        "sum"),
        )
    a = agg(apr).rename(columns={"AMT": "APR_AMT", "QTY": "APR_QTY"})
    m = agg(may).rename(columns={"AMT": "MAY_AMT", "QTY": "MAY_QTY"})
    df = pd.merge(a, m, on=group_col, how="outer").fillna(0)

    df["APR_AMT"] = (df["APR_AMT"] / CRORE).round(2)
    df["MAY_AMT"] = (df["MAY_AMT"] / CRORE).round(2)
    df["APR_QTY"] = df["APR_QTY"].astype(int)
    df["MAY_QTY"] = df["MAY_QTY"].astype(int)

    df["AMT_GRW"] = np.where(
        df["APR_AMT"] == 0, np.nan,
        ((df["MAY_AMT"] - df["APR_AMT"]) / df["APR_AMT"] * 100).round(2),
    )
    df["QTY_GRW"] = np.where(
        df["APR_QTY"] == 0, np.nan,
        ((df["MAY_QTY"] - df["APR_QTY"]) / df["APR_QTY"] * 100).round(2),
    )

    t_amt = df["MAY_AMT"].sum()
    t_qty = df["MAY_QTY"].sum()
    df["AMT_SHR"] = (df["MAY_AMT"] / t_amt * 100).round(2) if t_amt else 0
    df["QTY_SHR"] = (df["MAY_QTY"] / t_qty * 100).round(2) if t_qty else 0

    df = df.sort_values("MAY_AMT", ascending=False).reset_index(drop=True)
    return df

def add_grand_total(df, label_col):
    tot = {label_col: "Grand Total"}
    for c in ["APR_AMT", "MAY_AMT"]:
        tot[c] = round(df[c].sum(), 2)
    for c in ["APR_QTY", "MAY_QTY"]:
        tot[c] = int(df[c].sum())
    apr_a, may_a = tot["APR_AMT"], tot["MAY_AMT"]
    apr_q, may_q = tot["APR_QTY"], tot["MAY_QTY"]
    tot["AMT_GRW"] = round((may_a - apr_a) / apr_a * 100, 2) if apr_a else np.nan
    tot["QTY_GRW"] = round((may_q - apr_q) / apr_q * 100, 2) if apr_q else np.nan
    tot["AMT_SHR"] = 100.0
    tot["QTY_SHR"] = 100.0
    return pd.concat([df, pd.DataFrame([tot])], ignore_index=True)


COL_WIDTHS = [24, 12, 12, 11, 11, 2, 12, 12, 11, 11]

def _c(ws, row, col, value=None, font=None, fill=None,
       align=None, border=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if fmt:    cell.number_format = fmt
    return cell

def write_header(ws, row, label_name, month1_name, month2_name, sc=1):
    def hdr(r, c, v, colspan=1, rowspan=1):
        if colspan > 1 or rowspan > 1:
            ws.merge_cells(
                start_row=r, start_column=c,
                end_row=r + rowspan - 1, end_column=c + colspan - 1,
            )
        cell = ws.cell(row=r, column=c, value=v)
        cell.font   = HDR_FONT
        cell.fill   = WHITE_FILL
        cell.border = THIN_BDR
        cell.alignment = CENTER
        return cell

    hdr(row,   sc,    label_name,   rowspan=2)
    hdr(row,   sc+1,  month1_name.upper())
    hdr(row,   sc+2,  month2_name.upper())
    hdr(row,   sc+3,  "GROWTH %",   rowspan=2)
    hdr(row,   sc+4,  f"SHARE {month2_name.upper()}")
    
    ws.cell(row=row,   column=sc+5).border = Border()
    ws.cell(row=row+1, column=sc+5).border = Border()
    
    hdr(row,   sc+6,  month1_name.upper())
    hdr(row,   sc+7,  month2_name.upper())
    hdr(row,   sc+8,  "GROWTH %",   rowspan=2)
    hdr(row,   sc+9,  f"SHARE {month2_name.upper()}")

    hdr(row+1, sc+1, "AMT")
    hdr(row+1, sc+2, "AMT")
    hdr(row+1, sc+4, "AMT")
    hdr(row+1, sc+6, "QTY")
    hdr(row+1, sc+7, "QTY")
    hdr(row+1, sc+9, "QTY")

    return row + 2

def write_rows(ws, df, label_col, row, sc=1):
    for _, r in df.iterrows():
        is_tot = str(r[label_col]).strip().lower() == "grand total"
        fill  = TOT_FILL  if is_tot else WHITE_FILL
        font  = TOT_FONT  if is_tot else BODY_FONT

        def put(col, val, fmt=None, growth=False):
            v = None if (isinstance(val, float) and np.isnan(val)) else val
            cf = _c(ws, row, col, v, font=font, fill=fill,
                    align=RIGHT if col > sc else LEFT,
                    border=THIN_BDR, fmt=fmt)
            if growth and v is not None and not is_tot:
                cf.fill = NEG_FILL if v < 0 else WHITE_FILL
            return cf

        put(sc,    r[label_col])
        put(sc+1,  r["APR_AMT"], FMT_AMT)
        put(sc+2,  r["MAY_AMT"], FMT_AMT)
        put(sc+3,  r["AMT_GRW"], FMT_PCT, growth=True)
        put(sc+4,  r["AMT_SHR"], FMT_PCT)
        ws.cell(row=row, column=sc+5)
        put(sc+6,  r["APR_QTY"], FMT_QTY)
        put(sc+7,  r["MAY_QTY"], FMT_QTY)
        put(sc+8,  r["QTY_GRW"], FMT_PCT, growth=True)
        put(sc+9,  r["QTY_SHR"], FMT_PCT)
        row += 1
    return row

def set_col_widths(ws, sc=1):
    for i, w in enumerate(COL_WIDTHS, start=sc):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_table(ws, df, label_col, section_label, month1_name, month2_name, start_row=1, sc=1):
    row = write_header(ws, start_row, section_label, month1_name, month2_name, sc)
    row = write_rows(ws, df, label_col, row, sc)
    set_col_widths(ws, sc)
    ws.freeze_panes = ws.cell(row=start_row + 2, column=sc + 1)
    return row + 1

def to_pdf_rows(df, label_col, m1, m2):
    hdr = [label_col, f"{m1.upper()} AMT\n(Cr)", f"{m2.upper()} AMT\n(Cr)", "GRW %", "SHR %",
           f"{m1.upper()} QTY", f"{m2.upper()} QTY", "GRW %", "SHR %"]
    rows = [hdr]
    for _, r in df.iterrows():
        def fmt(v, typ):
            if isinstance(v, float) and np.isnan(v): return "-"
            if typ == "amt": return f"{v:,.2f}"
            if typ == "qty": return f"{int(v):,}"
            if typ == "pct": return f"{v:.2f}%"
            return str(v)
        rows.append([
            r[label_col],
            fmt(r["APR_AMT"], "amt"), fmt(r["MAY_AMT"], "amt"),
            fmt(r["AMT_GRW"], "pct"), fmt(r["AMT_SHR"], "pct"),
            fmt(r["APR_QTY"], "qty"), fmt(r["MAY_QTY"], "qty"),
            fmt(r["QTY_GRW"], "pct"), fmt(r["QTY_SHR"], "pct"),
        ])
    return rows

def build_pdf(sections, out_file_obj, title):
    doc = SimpleDocTemplate(out_file_obj, pagesize=landscape(A3),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    title_sty = ParagraphStyle("T", parent=styles["Heading1"],
                               fontName="Helvetica-Bold", fontSize=13,
                               textColor=colors.HexColor("#1F4E79"), spaceAfter=8)
    sec_sty   = ParagraphStyle("S", parent=styles["Heading2"],
                               fontName="Helvetica-Bold", fontSize=9,
                               textColor=colors.white,
                               backColor=colors.HexColor("#1F4E79"), spaceAfter=4)
    story = [Paragraph(title, title_sty), Spacer(1, 0.4*cm)]

    CW = [4.2*cm, 2.2*cm, 2.2*cm, 1.8*cm, 1.8*cm,
          2.0*cm, 2.0*cm, 1.8*cm, 1.8*cm]
    HDR_BG = colors.HexColor("#D9D9D9")
    TOT_BG  = colors.HexColor("#BDD7EE")
    NEG_BG  = colors.HexColor("#FFC7CE")
    ALT_BG  = colors.HexColor("#F9F9F9")

    for section_label, df, label_col, m1, m2 in sections:
        story.append(Paragraph(section_label, sec_sty))
        tbl_data = to_pdf_rows(df, label_col, m1, m2)
        n = len(tbl_data)
        tbl = Table(tbl_data, colWidths=CW, repeatRows=1)
        cmds = [
            ("BACKGROUND",  (0,0), (-1,0), HDR_BG),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 7),
            ("ALIGN",       (1,0), (-1,-1), "RIGHT"),
            ("ALIGN",       (0,0), (0,-1), "LEFT"),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0,1), (-1,n-2), [colors.white, ALT_BG]),
            ("BACKGROUND",  (0,n-1), (-1,n-1), TOT_BG),
            ("FONTNAME",    (0,n-1), (-1,n-1), "Helvetica-Bold"),
        ]
        for ri, r in enumerate(df.itertuples(), start=1):
            if hasattr(r, "AMT_GRW") and r.AMT_GRW is not None:
                g = r.AMT_GRW
                if not (isinstance(g, float) and np.isnan(g)) and g < 0:
                    cmds.append(("BACKGROUND", (3, ri), (3, ri), NEG_BG))
            if hasattr(r, "QTY_GRW") and r.QTY_GRW is not None:
                g = r.QTY_GRW
                if not (isinstance(g, float) and np.isnan(g)) and g < 0:
                    cmds.append(("BACKGROUND", (7, ri), (7, ri), NEG_BG))
        tbl.setStyle(TableStyle(cmds))
        story.append(tbl)
        story.append(Spacer(1, 0.5*cm))

    doc.build(story)

def generate_monthly_report_zip(y1, m1, y2, m2):
    import calendar
    month1_name = calendar.month_abbr[m1]
    month2_name = calendar.month_abbr[m2]
    
    apr_raw = load_from_db(y1, m1)
    may_raw = load_from_db(y2, m2)
    apr = add_pdt_cat(apr_raw)
    may = add_pdt_cat(may_raw)

    wb = Workbook()
    wb.remove(wb.active)

    pdf_secs = []

    def new_ws(name):
        return wb.create_sheet(name[:31])

    # 1. Product Category Summary
    ws = new_ws("1_Category Summary")
    a1 = apr[apr["PDT_CAT"] != "VAS"]
    m1 = may[may["PDT_CAT"] != "VAS"]
    df1 = pivot(a1, m1, "PDT_CAT")
    ord_map = {v: i for i, v in enumerate(PDT_CAT_ORDER)}
    df1 = df1.sort_values(
        ["PDT_CAT"], key=lambda s: s.map(lambda x: ord_map.get(x, 99))
    ).reset_index(drop=True)
    df1 = add_grand_total(df1, "PDT_CAT")
    write_table(ws, df1, "PDT_CAT", "PDT CAT", month1_name, month2_name)
    pdf_secs.append(("1. Product Category Summary", df1, "PDT_CAT", month1_name, month2_name))

    # 2. Mobile Brand Summary
    ws = new_ws("2_Mobile Brand")
    a2 = apr[apr["PDT_CAT"] == "MOBILE"]
    m2 = may[may["PDT_CAT"] == "MOBILE"]
    df2 = pivot(a2, m2, "Brand")
    df2 = add_grand_total(df2, "Brand")
    write_table(ws, df2, "Brand", "MOBILE", month1_name, month2_name)
    pdf_secs.append(("2. Mobile Brand Summary", df2, "Brand", month1_name, month2_name))

    # 3. Laptop Brand Summary
    ws = new_ws("3_Laptop Brand")
    a3 = apr[apr["PDT_CAT"] == "LAPTOP"]
    m3 = may[may["PDT_CAT"] == "LAPTOP"]
    df3 = pivot(a3, m3, "Brand")
    df3 = add_grand_total(df3, "Brand")
    write_table(ws, df3, "Brand", "LAPTOP", month1_name, month2_name)
    pdf_secs.append(("3. Laptop Brand Summary", df3, "Brand", month1_name, month2_name))

    # 4. ACC Brand Summary
    ws = new_ws("4_ACC Brand")
    a4 = apr[apr["PDT_CAT"] == "ACC"]
    m4 = may[may["PDT_CAT"] == "ACC"]
    df4 = pivot(a4, m4, "Brand")
    df4 = add_grand_total(df4, "Brand")
    write_table(ws, df4, "Brand", "ACC", month1_name, month2_name)
    pdf_secs.append(("4. ACC Brand Summary", df4, "Brand", month1_name, month2_name))

    # 5. ACC Sub-Category Summary
    ws = new_ws("5_ACC Category")
    df5 = pivot(a4, m4, "Product")
    df5 = add_grand_total(df5, "Product")
    write_table(ws, df5, "Product", "ACC", month1_name, month2_name)
    pdf_secs.append(("5. ACC Sub-Category Summary", df5, "Product", month1_name, month2_name))

    # 6. CE Sub-Category Summary
    ws = new_ws("6_CE Category")
    a6 = apr[apr["PDT_CAT"] == "CE"]
    m6 = may[may["PDT_CAT"] == "CE"]
    df6 = pivot(a6, m6, "Product")
    df6 = add_grand_total(df6, "Product")
    write_table(ws, df6, "Product", "CE", month1_name, month2_name)
    pdf_secs.append(("6. CE Sub-Category Summary", df6, "Product", month1_name, month2_name))

    # 7. CE Brand Summary
    ws = new_ws("7_CE Brand")
    df7 = pivot(a6, m6, "Brand")
    df7 = add_grand_total(df7, "Brand")
    write_table(ws, df7, "Brand", "CE", month1_name, month2_name)
    pdf_secs.append(("7. CE Brand Summary", df7, "Brand", month1_name, month2_name))

    # 8-11. CE Sub-Cat Brand Summaries
    subcats = [
        ("8_TV Brand",    "TV",               "8. TV Brand Summary"),
        ("9_Fridge Brand","REFRIGERATORS",    "9. Refrigerator Brand Summary"),
        ("10_WM Brand",   "WASHING MACHINES", "10. Washing Machine Brand Summary"),
        ("11_AC Brand",   "AIR CONDITIONER",  "11. AC Brand Summary"),
    ]
    for sname, subcat, pdflabel in subcats:
        ws = new_ws(sname)
        a_s = apr_raw[apr_raw["Product"].str.upper() == subcat]
        m_s = may_raw[may_raw["Product"].str.upper() == subcat]
        if a_s.empty and m_s.empty:
            ws.cell(1, 1, f"No data for {subcat}")
            continue
        df_s = pivot(a_s, m_s, "Brand")
        df_s = add_grand_total(df_s, "Brand")
        label = subcat.title()
        write_table(ws, df_s, "Brand", label, month1_name, month2_name)
        pdf_secs.append((pdflabel, df_s, "Brand", month1_name, month2_name))

    # 12. VAS Summary
    ws = new_ws("12_VAS Summary")
    a12 = apr[apr["PDT_CAT"] == "VAS"]
    m12 = may[may["PDT_CAT"] == "VAS"]
    if not (a12.empty and m12.empty):
        df12 = pivot(a12, m12, "Product")
        df12 = add_grand_total(df12, "Product")
        write_table(ws, df12, "Product", "VAS", month1_name, month2_name)
        pdf_secs.append(("12. Value Added Services", df12, "Product", month1_name, month2_name))

    # Write Excel to bytes
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    # Write PDF to bytes
    pdf_io = io.BytesIO()
    title = f"Monthly Category & Brand Performance Analysis — {month1_name.upper()} vs {month2_name.upper()}"
    build_pdf(pdf_secs, pdf_io, title)
    pdf_io.seek(0)

    # Package both into a ZIP
    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"Monthly_Analysis_{month1_name}_{month2_name}.xlsx", excel_io.getvalue())
        zf.writestr(f"Monthly_Analysis_{month1_name}_{month2_name}.pdf", pdf_io.getvalue())
    zip_io.seek(0)

    return zip_io
