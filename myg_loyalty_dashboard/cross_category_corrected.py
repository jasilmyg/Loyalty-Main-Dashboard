"""
CORRECTED Cross-Category Logic:
  Cross-sell = non-mobile item bought on:
    1. SAME INVOICE as a mobile purchase (bundled at POS)
    OR
    2. Any LATER DATE after first mobile purchase date

Mobile-Only = NONE of the above
"""
import os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'myg_loyalty_dashboard.settings')
sys.path.insert(0, '.')
django.setup()
from analytics.clickhouse_service import get_ch_client
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ch = get_ch_client()

EXCLUDE = """'STATIONERY ITEMS','SCHEME','GDOT CARE','D SPARE','OSG WARRANTY',
             'SERVICE','TOTAL SECURITY','LG AMC','SERVICE CHARGES',
             'DEMO','DEMO LAPTOP','DEMO ACCESSORIES','MYG DOMO','MYG VERSE',
             'DIY','CONTRACT WORK','CEGI','RIG','PROTECT MAX','CARE PLUS',
             'MOBILE ANTIVIRUS','HA ACCESSORIES','CCTV','MONITOR'"""

# ── Step 1: Corrected stats ──────────────────────────────────────────────────
print("=" * 70)
print("  Q13 CORRECTED — Mobile Cross-Category (same invoice OR later date)")
print("=" * 70)
print("  Fetching corrected numbers... (~3 min)")

total_mobile = ch.query("""
    SELECT countDistinct(ai.customer_mobile)
    FROM azure_invoice_report ai
    INNER JOIN azure_sales_report sr ON ai.invoice_no = sr.invoice_no
    INNER JOIN item_master m         ON sr.item_code = m.item_code
    WHERE m.product = 'MOBILE'
      AND length(trim(ai.customer_mobile)) >= 10
      AND toDate(ai.date) BETWEEN '2021-01-01' AND '2026-08-29'
      AND toDate(ai.date) != '1970-01-01'
""").result_rows[0][0]

# Cross-shoppers: same invoice OR later date
total_cross = ch.query(f"""
    WITH mobile_buyers AS (
        SELECT
            ai.customer_mobile               AS mobile,
            min(toDate(ai.date))             AS first_mobile_date,
            groupArray(ai.invoice_no)        AS mobile_invoices
        FROM azure_invoice_report ai
        INNER JOIN azure_sales_report sr ON ai.invoice_no = sr.invoice_no
        INNER JOIN item_master m         ON sr.item_code = m.item_code
        WHERE m.product = 'MOBILE'
          AND length(trim(ai.customer_mobile)) >= 10
          AND toDate(ai.date) BETWEEN '2021-01-01' AND '2026-08-29'
          AND toDate(ai.date) != '1970-01-01'
        GROUP BY ai.customer_mobile
    )
    SELECT countDistinct(ai2.customer_mobile)
    FROM azure_invoice_report ai2
    INNER JOIN azure_sales_report sr2 ON ai2.invoice_no = sr2.invoice_no
    INNER JOIN item_master m2         ON sr2.item_code = m2.item_code
    INNER JOIN mobile_buyers mb       ON ai2.customer_mobile = mb.mobile
    WHERE m2.product NOT IN ({EXCLUDE})
      AND m2.product != 'MOBILE'
      AND (
            has(mb.mobile_invoices, ai2.invoice_no)          -- same invoice as mobile
            OR toDate(ai2.date) > mb.first_mobile_date       -- or any later date
          )
      AND toDate(ai2.date) != '1970-01-01'
      AND length(trim(ai2.customer_mobile)) >= 10
""").result_rows[0][0]

mobile_only = total_mobile - total_cross

print(f"\n  Total Mobile Buyers (2021–2026)    : {total_mobile:>10,}")
print(f"  Cross-shoppers (corrected)         : {total_cross:>10,}  ({total_cross/total_mobile*100:.1f}%)")
print(f"  Mobile-Only (corrected)            : {mobile_only:>10,}  ({mobile_only/total_mobile*100:.1f}%)")

# ── Category breakdown ───────────────────────────────────────────────────────
print("\n  Fetching category breakdown...")

CAT_MAP = {
    'Glamshield / Screen Protection': ['GLAMSHIELD'],
    'Ear Wearables':                  ['EAR WEARABLES'],
    'Accessories (Mobile/General)':   ['ACCESSORIES','IT ACCESSORIES','LAPTOP BAG'],
    'Smart Watch':                    ['SMART WATCH'],
    'Audio / Home Theatre':           ['AUDIO','HOME THEATRE'],
    'Small Appliances / Kitchen HA':  ['SMALL APPLIANCES','HOME APPLIANCES','MICROWAVE OVEN','CROCKERY','DISH WASHER','HOUSE HOLD'],
    'TV':                             ['TV'],
    'Laptop / IT':                    ['LAPTOP','TABLET','PRINTER','DESKTOP','IT PRODUCT','STORAGE DEVICES','GAMING','CAMERA'],
    'AC':                             ['AIR CONDITIONER','AC OUTDOOR','STABILIZER'],
    'Washing Machine':                ['WASHING MACHINES','DRYER'],
    'Refrigerator':                   ['REFRIGERATORS','FREEZER'],
    'Personal Care / Fragrance':      ['PERSONAL CARE','FRAGRANCE'],
    'Gift Items':                     ['GIFT ITEMS'],
    'Smart Choice / Offers':          ['SMART CHOICE','OFFER KIT'],
}

raw = ch.query(f"""
    WITH mobile_buyers AS (
        SELECT
            ai.customer_mobile               AS mobile,
            min(toDate(ai.date))             AS first_mobile_date,
            groupArray(ai.invoice_no)        AS mobile_invoices
        FROM azure_invoice_report ai
        INNER JOIN azure_sales_report sr ON ai.invoice_no = sr.invoice_no
        INNER JOIN item_master m         ON sr.item_code = m.item_code
        WHERE m.product = 'MOBILE'
          AND length(trim(ai.customer_mobile)) >= 10
          AND toDate(ai.date) BETWEEN '2021-01-01' AND '2026-08-29'
          AND toDate(ai.date) != '1970-01-01'
        GROUP BY ai.customer_mobile
    )
    SELECT
        m2.product,
        countDistinct(ai2.customer_mobile) AS buyers
    FROM azure_invoice_report ai2
    INNER JOIN azure_sales_report sr2 ON ai2.invoice_no = sr2.invoice_no
    INNER JOIN item_master m2         ON sr2.item_code = m2.item_code
    INNER JOIN mobile_buyers mb       ON ai2.customer_mobile = mb.mobile
    WHERE m2.product NOT IN ({EXCLUDE})
      AND m2.product != 'MOBILE'
      AND (
            has(mb.mobile_invoices, ai2.invoice_no)
            OR toDate(ai2.date) > mb.first_mobile_date
          )
      AND toDate(ai2.date) != '1970-01-01'
      AND length(trim(ai2.customer_mobile)) >= 10
    GROUP BY m2.product
    ORDER BY buyers DESC
""").result_rows

prod_buyers = {str(r[0]): int(r[1]) for r in raw}
cat_results = {}
for cat, prods in CAT_MAP.items():
    cat_results[cat] = sum(prod_buyers.get(p, 0) for p in prods)

print(f"\n  {'Business Category':<38} {'Buyers':>10} {'% of Mobile':>14} {'% of Cross':>12}")
print("  " + "-" * 80)
for cat, buyers in sorted(cat_results.items(), key=lambda x: x[1], reverse=True):
    if buyers == 0: continue
    print(f"  {cat:<38} {buyers:>10,} {buyers/total_mobile*100:>13.1f}% {buyers/total_cross*100:>11.1f}%")
print("=" * 70)

# ── Step 2: Export corrected mobile-only customers ───────────────────────────
print("\n  Fetching corrected mobile-only customer list... (~5 min)")

rows = ch.query(f"""
    WITH mobile_buyers AS (
        SELECT
            ai.customer_mobile               AS mobile,
            any(ai.customer_type)            AS cust_type,
            min(toDate(ai.date))             AS first_date,
            max(toDate(ai.date))             AS last_date,
            countDistinct(ai.invoice_no)     AS total_invoices,
            sum(toFloat64(sr.sold_price))    AS total_spend,
            any(ai.branch)                   AS branch_code,
            any(ai.rbm)                      AS rbm,
            any(ai.bdm)                      AS bdm,
            groupArray(ai.invoice_no)        AS mobile_invoices
        FROM azure_invoice_report ai
        INNER JOIN azure_sales_report sr ON ai.invoice_no = sr.invoice_no
        INNER JOIN item_master m         ON sr.item_code = m.item_code
        WHERE m.product = 'MOBILE'
          AND length(trim(ai.customer_mobile)) >= 10
          AND toDate(ai.date) BETWEEN '2021-01-01' AND '2026-08-29'
          AND toDate(ai.date) != '1970-01-01'
        GROUP BY ai.customer_mobile
    ),
    cross_shoppers AS (
        SELECT DISTINCT ai2.customer_mobile AS mobile
        FROM azure_invoice_report ai2
        INNER JOIN azure_sales_report sr2 ON ai2.invoice_no = sr2.invoice_no
        INNER JOIN item_master m2         ON sr2.item_code = m2.item_code
        INNER JOIN mobile_buyers mb       ON ai2.customer_mobile = mb.mobile
        WHERE m2.product NOT IN ({EXCLUDE})
          AND m2.product != 'MOBILE'
          AND (
                has(mb.mobile_invoices, ai2.invoice_no)
                OR toDate(ai2.date) > mb.first_date
              )
          AND toDate(ai2.date) != '1970-01-01'
          AND length(trim(ai2.customer_mobile)) >= 10
    )
    SELECT
        mb.mobile,
        mb.cust_type,
        mb.first_date,
        mb.last_date,
        mb.total_invoices,
        mb.total_spend,
        mb.branch_code,
        b.branch_name,
        mb.rbm,
        mb.bdm
    FROM mobile_buyers mb
    LEFT JOIN branch_master b ON mb.branch_code = b.code
    WHERE mb.mobile NOT IN (SELECT mobile FROM cross_shoppers)
    ORDER BY mb.total_spend DESC
""").result_rows

print(f"  → Corrected mobile-only: {len(rows):,}")

df = pd.DataFrame(rows, columns=[
    'Mobile','Customer Type','First Purchase Date','Last Purchase Date',
    'Total Invoices','Total Mobile Spend (Rs)','Branch Code','Branch Name','RBM','BDM'
])
df['First Purchase Date']  = pd.to_datetime(df['First Purchase Date']).dt.date
df['Last Purchase Date']   = pd.to_datetime(df['Last Purchase Date']).dt.date
df['Total Mobile Spend (Rs)'] = df['Total Mobile Spend (Rs)'].round(0).astype(int)
df['Days Since Last Purchase'] = (
    pd.to_datetime('2026-08-29') - pd.to_datetime(df['Last Purchase Date'])
).dt.days

# Save CSV
CSV_FILE = r'C:\Users\jasil_myg\Downloads\Mobile_Only_Customers_CORRECTED.csv'
df.drop(columns=['Branch Code']).to_csv(CSV_FILE, index=False, encoding='utf-8-sig')
print(f"  ✅ CSV saved → {CSV_FILE}")

# Save Summary Excel
OUTFILE = r'C:\Users\jasil_myg\Downloads\Mobile_Only_Customers_CORRECTED_SUMMARY.xlsx'
wb = Workbook()
HDR_FILL = PatternFill('solid', fgColor='1E3A5F')
HDR_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
ALT_FILL = PatternFill('solid', fgColor='EEF4FB')
BODY_FONT = Font(name='Calibri', size=9)
THIN   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT_A = Alignment(horizontal='left', vertical='center')
RIGHT_A= Alignment(horizontal='right', vertical='center')

def write_sheet(ws, title, hdrs_widths, data_rows):
    ws.title = title
    ws.merge_cells(f'A1:{get_column_letter(len(hdrs_widths))}1')
    ws['A1'] = f'  myG — Mobile-Only (Corrected): {title}'
    ws['A1'].font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    ws['A1'].fill = PatternFill('solid', fgColor='1E3A5F')
    ws['A1'].alignment = LEFT_A
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 26
    for ci, (h, w) in enumerate(hdrs_widths, 1):
        c = ws.cell(2, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(data_rows, 3):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font = BODY_FONT; c.border = BORDER
            if fill: c.fill = fill
            c.alignment = RIGHT_A if isinstance(val, (int, float)) else LEFT_A
        ws.row_dimensions[ri].height = 15
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(hdrs_widths))}{2+len(data_rows)}'

# Overview
ws0 = wb.active; ws0.title = 'Overview'
ws0['A1'] = 'myG — Mobile-Only Customers (Corrected Logic)'
ws0['A1'].font = Font(name='Calibri', bold=True, size=14, color='1E3A5F')
ws0.row_dimensions[1].height = 28
stats = [
    ('Logic Applied', 'Same invoice as mobile OR any later date'),
    ('Total Mobile-Only Customers', f"{len(df):,}"),
    ('Total Mobile Spend (Rs)', f"Rs {df['Total Mobile Spend (Rs)'].sum():,.0f}"),
    ('Avg Spend per Customer', f"Rs {df['Total Mobile Spend (Rs)'].mean():,.0f}"),
    ('Avg Days Since Last Purchase', f"{df['Days Since Last Purchase'].mean():.0f} days"),
    ('Inactive >365 days', f"{(df['Days Since Last Purchase']>365).sum():,}  ({(df['Days Since Last Purchase']>365).mean()*100:.1f}%)"),
    ('Inactive <90 days', f"{(df['Days Since Last Purchase']<90).sum():,}  ({(df['Days Since Last Purchase']<90).mean()*100:.1f}%)"),
    ('Report Date', '29-Aug-2026'),
    ('Full Data', 'Mobile_Only_Customers_CORRECTED.csv (Downloads)'),
]
for i, (k, v) in enumerate(stats, 2):
    ws0.cell(i, 1, k).font = Font(name='Calibri', bold=True, size=10, color='1E3A5F')
    ws0.cell(i, 2, v).font = Font(name='Calibri', size=10)
    ws0.column_dimensions['A'].width = 42
    ws0.column_dimensions['B'].width = 48

# By Branch
bg = df.groupby('Branch Name').agg(
    Customers=('Mobile','count'),
    Total_Spend=('Total Mobile Spend (Rs)','sum'),
    Avg_Spend=('Total Mobile Spend (Rs)','mean'),
    Avg_Days=('Days Since Last Purchase','mean'),
).reset_index().sort_values('Customers', ascending=False)
bg['Pct'] = bg['Customers'] / len(df) * 100
ws1 = wb.create_sheet()
write_sheet(ws1, 'By Branch', [
    ('Branch Name',24),('Customers',12),('% of Total',11),
    ('Total Spend (Rs)',18),('Avg Spend/Cust',16),('Avg Days Inactive',16)
], [(r['Branch Name'],int(r['Customers']),round(r['Pct'],2),
     round(r['Total_Spend'],0),round(r['Avg_Spend'],0),round(r['Avg_Days'],0))
    for _,r in bg.iterrows()])

# By RBM
rg = df.groupby('RBM').agg(
    Customers=('Mobile','count'),
    Total_Spend=('Total Mobile Spend (Rs)','sum'),
    Avg_Spend=('Total Mobile Spend (Rs)','mean'),
    Avg_Days=('Days Since Last Purchase','mean'),
).reset_index().sort_values('Customers', ascending=False)
rg['Pct'] = rg['Customers'] / len(df) * 100
ws2 = wb.create_sheet()
write_sheet(ws2, 'By RBM', [
    ('RBM',22),('Customers',12),('% of Total',11),
    ('Total Spend (Rs)',18),('Avg Spend/Cust',16),('Avg Days Inactive',16)
], [(r['RBM'],int(r['Customers']),round(r['Pct'],2),
     round(r['Total_Spend'],0),round(r['Avg_Spend'],0),round(r['Avg_Days'],0))
    for _,r in rg.iterrows()])

# Inactivity Buckets
ws3 = wb.create_sheet('Inactivity Buckets')
ws3['A1'] = 'myG — Mobile-Only Customers (Corrected): Inactivity Segments'
ws3['A1'].font = Font(name='Calibri', bold=True, size=12, color='1E3A5F')
ws3.row_dimensions[1].height = 24
buckets = [
    ('< 30 days',    df['Days Since Last Purchase'].lt(30).sum()),
    ('30-90 days',   df['Days Since Last Purchase'].between(30,90).sum()),
    ('91-180 days',  df['Days Since Last Purchase'].between(91,180).sum()),
    ('181-365 days', df['Days Since Last Purchase'].between(181,365).sum()),
    ('1-2 years',    df['Days Since Last Purchase'].between(366,730).sum()),
    ('2-3 years',    df['Days Since Last Purchase'].between(731,1095).sum()),
    ('> 3 years',    df['Days Since Last Purchase'].gt(1095).sum()),
]
pmap = {
    '< 30 days':'Recently active - nurture',
    '30-90 days':'Warm - send cross-sell offer now',
    '91-180 days':'At risk - send reactivation',
    '181-365 days':'Cold - discount offer needed',
    '1-2 years':'Lapsed - strong incentive needed',
    '2-3 years':'Highly lapsed - last-chance campaign',
    '> 3 years':'Likely churned permanently',
}
hdrs = [('Inactivity Bucket',22),('Customers',14),('% of Total',12),('Action Priority',32)]
for ci,(h,w) in enumerate(hdrs,1):
    c = ws3.cell(2,ci,h); c.font=HDR_FONT; c.fill=HDR_FILL; c.border=BORDER
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws3.column_dimensions[get_column_letter(ci)].width=w
ws3.row_dimensions[2].height=22
total=len(df)
for ri,(label,count) in enumerate(buckets,3):
    fill=ALT_FILL if ri%2==0 else None
    for ci,val in enumerate([label,count,round(count/total*100,1),pmap[label]],1):
        c=ws3.cell(ri,ci,val); c.font=BODY_FONT; c.border=BORDER
        if fill: c.fill=fill
        c.alignment=LEFT_A
    ws3.row_dimensions[ri].height=16

wb.save(OUTFILE)
print(f"  ✅ Excel saved → {OUTFILE}")
print(f"\n  Summary:")
print(f"    Previous mobile-only count : 1,726,467")
print(f"    Corrected mobile-only count: {len(df):,}")
print(f"    Customers now correctly moved to cross-sell: {1726467 - len(df):,}")
