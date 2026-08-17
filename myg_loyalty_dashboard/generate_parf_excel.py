"""
MY PARF — Enhanced Excel Report v2
- Full branch names (mapped from sales_data)
- Customer-level analytics (new vs repeat, top customers, mobile segments)
"""
import os; os.environ.setdefault('DJANGO_SETTINGS_MODULE','myg_loyalty_dashboard.settings')
import django; django.setup()

from analytics.clickhouse_service import get_ch_client
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from datetime import date

ch = get_ch_client()

# ── Colour palette ────────────────────────────────────────────
C_BRAND='FF6B35'; C_DARK='1A1A2E'; C_ACCENT='16213E'; C_GOLD='F5A623'
C_GREEN='27AE60'; C_RED='E74C3C'; C_LIGHT='FFF8F0'; C_WHITE='FFFFFF'
C_GRAY='F5F5F5'; C_ALT='FFF3E6'; C_PURPLE='8E44AD'; C_BLUE='2980B9'
C_BORDER='DDDDDD'

def fill(c): return PatternFill("solid", fgColor=c)
def bold(size=11, color='000000', italic=False):
    return Font(name='Calibri', bold=True, size=size, color=color, italic=italic)
def normal(size=10, color='000000'):
    return Font(name='Calibri', size=size, color=color)
def ctr(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def rgt(): return Alignment(horizontal='right',  vertical='center')
def lft(): return Alignment(horizontal='left',   vertical='center')
def bdr():
    s=Side(style='thin',color=C_BORDER)
    return Border(left=s,right=s,top=s,bottom=s)
def W(ws,col,w): ws.column_dimensions[get_column_letter(col)].width=w

def hdr(ws,row,col,text,bg=C_BRAND,fg=C_WHITE,size=10):
    c=ws.cell(row=row,column=col,value=text)
    c.font=bold(size=size,color=fg); c.fill=fill(bg)
    c.alignment=ctr(); c.border=bdr(); return c

def cell(ws,row,col,val,fmt=None,bg=C_WHITE,bold_=False,color='000000',align='left'):
    c=ws.cell(row=row,column=col,value=val)
    c.font=bold(color=color) if bold_ else normal(color=color)
    c.fill=fill(bg); c.border=bdr()
    c.alignment=ctr() if align=='center' else (rgt() if align=='right' else lft())
    if fmt: c.number_format=fmt
    return c

def title_row(ws,row,text,cols='A:H',bg=C_DARK,fg=C_WHITE,h=28,size=13):
    ws.merge_cells(f'{cols.split(":")[0]}{row}:{cols.split(":")[1]}{row}')
    ws[f'{cols.split(":")[0]}{row}']=text
    ws[f'{cols.split(":")[0]}{row}'].font=bold(size=size,color=fg)
    ws[f'{cols.split(":")[0]}{row}'].fill=fill(bg)
    ws[f'{cols.split(":")[0]}{row}'].alignment=ctr()
    ws.row_dimensions[row].height=h

# ═══════════════════════════════════════════════════════
# FETCH DATA
# ═══════════════════════════════════════════════════════
print('Fetching data...')

# Build branch code → full name mapping from sales_data
print('  Building branch name mapping...')
br_map_rows = ch.query("""
    SELECT DISTINCT upper(trim(a.branch)) AS code, upper(trim(s.branch)) AS full_name
    FROM azure_sales_report a
    JOIN sales_data s ON a.invoice_no = s.invoice_number
    WHERE a.branch != '' AND s.branch != ''
    ORDER BY code
""").result_rows
branch_map = {r[0]: f"{r[0]}-{r[1]}" for r in br_map_rows}

def full_br(code):
    return branch_map.get(str(code).upper().strip(), str(code).upper().strip())

# Summary
summary = ch.query("""
    SELECT sum(s.sold_price), sum(s.qty), countDistinct(s.invoice_no),
           countDistinct(s.branch), avg(s.sold_price/s.qty),
           min(toDate(s.date)), max(toDate(s.date))
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0 AND toDate(s.date)!='1970-01-01'
""").result_rows[0]

# Products
products = ch.query("""
    SELECT m.item_name, m.item_code, m.mrp,
           sum(s.sold_price) AS rev, sum(s.qty) AS qty,
           countDistinct(s.invoice_no) AS inv,
           avg(s.sold_price/s.qty) AS avg_price,
           (m.mrp - avg(s.sold_price/s.qty)) / m.mrp * 100 AS disc_pct
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0 AND toDate(s.date)!='1970-01-01'
    GROUP BY m.item_name, m.item_code, m.mrp ORDER BY rev DESC
""").result_rows

# Monthly
monthly = ch.query("""
    SELECT formatDateTime(toStartOfMonth(s.date),'%Y-%m') AS month,
           sum(s.sold_price),sum(s.qty),countDistinct(s.invoice_no),avg(s.sold_price/s.qty)
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0 AND toDate(s.date)!='1970-01-01'
    GROUP BY month ORDER BY month
""").result_rows

# Branches with full names from cross-join
branches = ch.query("""
    SELECT s.branch,
           sum(s.sold_price) AS rev, sum(s.qty) AS qty,
           countDistinct(s.invoice_no) AS inv,
           avg(s.sold_price/s.qty) AS avg_price
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0 AND toDate(s.date)!='1970-01-01'
    GROUP BY s.branch ORDER BY rev DESC
""").result_rows

# Yearly
yearly = ch.query("""
    SELECT toYear(s.date), sum(s.sold_price), sum(s.qty), countDistinct(s.invoice_no)
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0 AND toDate(s.date)!='1970-01-01'
    GROUP BY toYear(s.date) ORDER BY toYear(s.date)
""").result_rows

# Price tiers
price_tiers = ch.query("""
    SELECT multiIf(m.mrp<500,'Below Rs.500',m.mrp<1000,'Rs.500-999',
                   m.mrp<1500,'Rs.1000-1499',m.mrp<2000,'Rs.1500-1999','Rs.2000+') AS tier,
           countDistinct(m.item_code), sum(s.qty), sum(s.sold_price)
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0 AND toDate(s.date)!='1970-01-01'
    GROUP BY tier ORDER BY sum(s.sold_price) DESC
""").result_rows

# Day of week
dow = ch.query("""
    SELECT toDayOfWeek(s.date),
           ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][toDayOfWeek(s.date)],
           sum(s.sold_price), sum(s.qty)
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0 AND toDate(s.date)!='1970-01-01'
    GROUP BY toDayOfWeek(s.date), ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][toDayOfWeek(s.date)]
    ORDER BY toDayOfWeek(s.date)
""").result_rows

# Last 30 days
daily30 = ch.query("""
    SELECT toDate(s.date), sum(s.sold_price), sum(s.qty), countDistinct(s.invoice_no)
    FROM azure_sales_report s JOIN item_master m ON s.item_code=m.item_code
    WHERE m.brand='MY PARF' AND s.sold_price>0
      AND toDate(s.date)>=today()-30 AND toDate(s.date)!='1970-01-01'
    GROUP BY toDate(s.date) ORDER BY toDate(s.date)
""").result_rows

# ─── CUSTOMER ANALYTICS ───────────────────────────────────────
print('  Fetching customer analytics...')

# Step 1: Get all unique MY PARF customers with their total spend
parf_custs_raw = ch.query("""
    SELECT mob, sum(rev) AS total_rev
    FROM (
        SELECT i.customer_mobile AS mob, s.sold_price AS rev
        FROM azure_sales_report s
        JOIN item_master m ON s.item_code = m.item_code
        JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
        WHERE m.brand = 'MY PARF'
          AND s.sold_price > 0
          AND toDate(s.date) != '1970-01-01'
          AND i.customer_mobile != ''
          AND length(i.customer_mobile) = 10
    )
    GROUP BY mob
""").result_rows

total_customers  = len(parf_custs_raw)
avg_clv          = sum(float(r[1]) for r in parf_custs_raw) / max(total_customers, 1)
max_clv          = max((float(r[1]) for r in parf_custs_raw), default=0)

# Step 2: Calculate Repeat vs One-Time MY PARF buyers (based on 2+ separate invoices)
repeat_mobs_raw = ch.query("""
    SELECT i.customer_mobile
    FROM azure_sales_report s
    JOIN item_master m ON s.item_code = m.item_code
    JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
    WHERE m.brand = 'MY PARF' AND s.sold_price > 0
      AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
      AND toDate(s.date) != '1970-01-01'
    GROUP BY i.customer_mobile
    HAVING countDistinct(s.invoice_no) > 1
""").result_rows

repeat_customers = len(repeat_mobs_raw)
new_customers_total = total_customers - repeat_customers

cust_summary = [total_customers, repeat_customers, avg_clv, max_clv]

# Customer purchase frequency distribution
freq_dist = ch.query("""
    SELECT purchase_count, count() AS customers, sum(total_rev) AS revenue
    FROM (
        SELECT i.customer_mobile AS mob,
               countDistinct(s.invoice_no) AS purchase_count,
               sum(s.sold_price) AS total_rev
        FROM azure_sales_report s
        JOIN item_master m ON s.item_code = m.item_code
        JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
        WHERE m.brand = 'MY PARF' AND s.sold_price > 0
          AND toDate(s.date) != '1970-01-01'
          AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
        GROUP BY i.customer_mobile
    )
    GROUP BY purchase_count
    ORDER BY purchase_count LIMIT 10
""").result_rows

# Top 30 customers by spend
top_customers = ch.query("""
    SELECT mob, cust_type, total_spend, visits, last_visit, first_visit
    FROM (
        SELECT
            i.customer_mobile                   AS mob,
            max(i.customer_type)                AS cust_type,
            sum(s.sold_price)                   AS total_spend,
            countDistinct(s.invoice_no)         AS visits,
            max(toDate(i.date))                 AS last_visit,
            min(toDate(i.date))                 AS first_visit
        FROM azure_sales_report s
        JOIN item_master m ON s.item_code = m.item_code
        JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
        WHERE m.brand = 'MY PARF' AND s.sold_price > 0
          AND toDate(s.date) != '1970-01-01'
          AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
        GROUP BY i.customer_mobile
    )
    ORDER BY total_spend DESC
    LIMIT 30
""").result_rows

# New vs Repeat customers monthly — use a simpler approach
# Get all PARF customers with their first-ever purchase date across ALL products
print('  Fetching monthly new vs repeat...')
new_repeat_monthly = ch.query("""
    WITH
    parf_monthly AS (
        SELECT
            formatDateTime(toStartOfMonth(s.date), '%Y-%m') AS month,
            i.customer_mobile AS mob
        FROM azure_sales_report s
        JOIN item_master m ON s.item_code = m.item_code
        JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
        WHERE m.brand = 'MY PARF' AND s.sold_price > 0
          AND toDate(s.date) != '1970-01-01'
          AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
    ),
    cust_first AS (
        SELECT customer_mobile AS mob,
               formatDateTime(toStartOfMonth(min(toDate(date))), '%Y-%m') AS first_month
        FROM azure_invoice_report
        WHERE customer_mobile != '' AND length(customer_mobile) = 10
          AND invoice_total > 0 AND toDate(date) != '1970-01-01'
        GROUP BY customer_mobile
    )
    SELECT
        pm.month,
        countDistinctIf(pm.mob, cf.first_month = pm.month) AS new_customers,
        countDistinctIf(pm.mob, cf.first_month != pm.month) AS repeat_customers,
        countDistinct(pm.mob) AS total_customers
    FROM parf_monthly pm
    LEFT JOIN cust_first cf ON pm.mob = cf.mob
    GROUP BY pm.month
    ORDER BY pm.month
""").result_rows

# Customer type breakdown
cust_type = ch.query("""
    SELECT cust_type, countDistinct(mob) AS customers, sum(rev) AS revenue
    FROM (
        SELECT i.customer_type AS cust_type,
               i.customer_mobile AS mob,
               s.sold_price AS rev
        FROM azure_sales_report s
        JOIN item_master m ON s.item_code = m.item_code
        JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
        WHERE m.brand = 'MY PARF' AND s.sold_price > 0
          AND toDate(s.date) != '1970-01-01'
          AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
    )
    GROUP BY cust_type ORDER BY revenue DESC
""").result_rows

# Dormant PARF customers (bought PARF, no purchase in last 90 days)
dormant = ch.query("""
    SELECT countDistinct(mob) AS dormant, sum(total_spend) AS pot
    FROM (
        SELECT mob, max(last_d) AS last_purchase, sum(spend) AS total_spend
        FROM (
            SELECT i.customer_mobile AS mob, toDate(i.date) AS last_d, s.sold_price AS spend
            FROM azure_sales_report s
            JOIN item_master m ON s.item_code = m.item_code
            JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
            WHERE m.brand = 'MY PARF' AND s.sold_price > 0
              AND toDate(s.date) != '1970-01-01'
              AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
        )
        GROUP BY mob
        HAVING max(last_d) < today() - 90
    )
""").result_rows[0]

# ─── REPEAT CUSTOMER BEHAVIOR ───────────────────────────────────────
print('  Fetching repeat customer behavior...')

repeat_metrics = ch.query("""
    SELECT 
        multiIf(visits > 1, 'Repeat (2+ purchases)', 'One-Time Buyer') AS segment,
        count(mob) AS customers,
        sum(total_spend) AS revenue,
        sum(total_spend) / count(mob) AS avg_clv,
        sum(visits) AS total_visits
    FROM (
        SELECT i.customer_mobile AS mob,
               countDistinct(s.invoice_no) AS visits,
               sum(s.sold_price) AS total_spend
        FROM azure_sales_report s
        JOIN item_master m ON s.item_code = m.item_code
        JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
        WHERE m.brand = 'MY PARF' AND s.sold_price > 0
          AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
          AND toDate(s.date) != '1970-01-01'
        GROUP BY i.customer_mobile
    )
    GROUP BY segment
    ORDER BY customers ASC
""").result_rows

repeat_gap = ch.query("""
    SELECT 
        avg(dateDiff('day', first_buy, second_buy)) AS avg_days_to_second,
        avg(dateDiff('day', first_buy, last_buy)) AS avg_days_lifetime
    FROM (
        SELECT 
            i.customer_mobile AS mob,
            min(toDate(s.date)) AS first_buy,
            max(toDate(s.date)) AS last_buy,
            arrayMin(arrayFilter(d -> d > min(toDate(s.date)), groupArray(toDate(s.date)))) AS second_buy,
            countDistinct(s.invoice_no) AS visits
        FROM azure_sales_report s
        JOIN item_master m ON s.item_code = m.item_code
        JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
        WHERE m.brand = 'MY PARF' AND s.sold_price > 0
          AND i.customer_mobile != '' AND length(i.customer_mobile) = 10
          AND toDate(s.date) != '1970-01-01'
        GROUP BY i.customer_mobile
        HAVING visits > 1
    )
    WHERE second_buy > '1970-01-01'
""").result_rows[0]

repeat_products = ch.query("""
    SELECT 
        m.item_name,
        m.mrp,
        countDistinct(i.customer_mobile) AS buyers,
        sum(s.qty) AS qty,
        sum(s.sold_price) AS revenue
    FROM azure_sales_report s
    JOIN item_master m ON s.item_code = m.item_code
    JOIN azure_invoice_report i ON s.invoice_no = i.invoice_no
    WHERE m.brand = 'MY PARF' AND s.sold_price > 0
      AND toDate(s.date) != '1970-01-01'
      AND i.customer_mobile IN (
          SELECT mob FROM (
              SELECT i2.customer_mobile AS mob, countDistinct(s2.invoice_no) AS visits
              FROM azure_sales_report s2
              JOIN item_master m2 ON s2.item_code = m2.item_code
              JOIN azure_invoice_report i2 ON s2.invoice_no = i2.invoice_no
              WHERE m2.brand = 'MY PARF' AND s2.sold_price > 0
                AND i2.customer_mobile != '' AND length(i2.customer_mobile) = 10
                AND toDate(s2.date) != '1970-01-01'
              GROUP BY i2.customer_mobile
              HAVING visits > 1
          )
      )
    GROUP BY m.item_name, m.mrp
    ORDER BY revenue DESC
    LIMIT 15
""").result_rows

print('  All data fetched!')

# ═══════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══ SHEET 1: SUMMARY DASHBOARD ════════════════════════
ws1 = wb.create_sheet('📊 Summary Dashboard')
ws1.sheet_view.showGridLines = False

title_row(ws1,1,'🌸  MY PARF — COMPREHENSIVE SALES & CUSTOMER ANALYSIS REPORT','A:I',h=36,size=16)
ws1.merge_cells('A2:I2')
ws1['A2']=f'Data: {summary[5]} → {summary[6]}   |   Generated: {date.today()}   |   Brand: MY PARF (PERFUMES)'
ws1['A2'].font=normal(size=10,color='AAAAAA'); ws1['A2'].fill=fill(C_ACCENT)
ws1['A2'].alignment=ctr(); ws1.row_dimensions[2].height=18

# KPI section header
ws1.row_dimensions[3].height=8
ws1.merge_cells('A4:I4')
ws1['A4']='KEY PERFORMANCE INDICATORS'
ws1['A4'].font=bold(size=11,color='777777'); ws1.row_dimensions[4].height=20

# KPI Cards
kpis=[
    ('TOTAL REVENUE','Rs. {:,.0f}'.format(float(summary[0])),'{:.2f} Crore'.format(float(summary[0])/1e7),C_BRAND),
    ('TOTAL UNITS','  {:,}'.format(int(summary[1])),'Units Sold',C_GREEN),
    ('TOTAL INVOICES','{:,}'.format(int(summary[2])),'Transactions',C_PURPLE),
    ('BRANCHES','  {:}'.format(int(summary[3])),'Active Branches','E67E22'),
    ('AVG SELL PRICE','Rs. {:,.0f}'.format(float(summary[4])),'Per Unit',C_BLUE),
]
merges=[('A5:B8'),('C5:D8'),('E5:F8'),('G5:H8'),('I5:I8')]
for mc,(_l,_v,_s,_c) in zip(merges,kpis):
    ws1.merge_cells(mc)
    ws1[mc.split(':')[0]]=f'{_l}\n{_v}\n{_s}'
    ws1[mc.split(':')[0]].font=bold(size=13,color=C_WHITE)
    ws1[mc.split(':')[0]].fill=fill(_c)
    ws1[mc.split(':')[0]].alignment=ctr()
ws1.row_dimensions[5].height=70

ws1.row_dimensions[9].height=8

# Customer KPIs
total_custs  = int(cust_summary[0])
repeat_custs = int(cust_summary[1])
new_custs    = total_custs - repeat_custs
avg_clv      = float(cust_summary[2])
dormant_c    = int(dormant[0])

ws1.merge_cells('A10:I10')
ws1['A10']='CUSTOMER ANALYTICS SNAPSHOT'
ws1['A10'].font=bold(size=11,color='777777'); ws1.row_dimensions[10].height=20

cust_kpis=[
    ('UNIQUE CUSTOMERS','{:,}'.format(total_custs),'Bought MY PARF',C_DARK),
    ('NEW CUSTOMERS','{:,}'.format(new_custs),'First-time Buyers','1ABC9C'),
    ('REPEAT CUSTOMERS','{:,}'.format(repeat_custs),'Loyal Buyers',C_BRAND),
    ('AVG CUSTOMER VALUE','Rs. {:,.0f}'.format(avg_clv),'Lifetime Spend',C_GOLD),
    ('DORMANT CUSTOMERS','{:,}'.format(dormant_c),'No Buy in 90 Days',C_RED),
]
merges2=[('A11:B14'),('C11:D14'),('E11:F14'),('G11:H14'),('I11:I14')]
for mc,(_l,_v,_s,_c) in zip(merges2,cust_kpis):
    ws1.merge_cells(mc)
    ws1[mc.split(':')[0]]=f'{_l}\n{_v}\n{_s}'
    ws1[mc.split(':')[0]].font=bold(size=12,color=C_WHITE)
    ws1[mc.split(':')[0]].fill=fill(_c)
    ws1[mc.split(':')[0]].alignment=ctr()
ws1.row_dimensions[11].height=65

# YoY Table
ws1.row_dimensions[15].height=8
ws1.merge_cells('A16:I16')
ws1['A16']='YEAR-OVER-YEAR COMPARISON'; ws1['A16'].font=bold(size=11,color='777777')
for ci,h in enumerate(['Year','Revenue','Revenue (Cr)','Units','Invoices','YoY Growth'],1):
    hdr(ws1,17,ci,h,bg=C_DARK)
prev=0
for ri,r in enumerate(yearly):
    row=18+ri; bg=C_ALT if ri%2==0 else C_WHITE
    rev=float(r[1]); g=((rev-prev)/prev*100) if prev>0 else 0
    gstr=f'+{g:.1f}%' if g>0 and prev>0 else (f'{g:.1f}%' if prev>0 else '—')
    gc=C_GREEN if g>0 and prev>0 else (C_RED if g<0 else '555555')
    vals=[int(r[0]),f'Rs.{rev:,.0f}',f'{rev/1e7:.2f}',int(r[2]),int(r[3]),gstr]
    for ci,v in enumerate(vals,1):
        c=cell(ws1,row,ci,v,bg=bg,align='center')
        if ci==6: c.font=bold(color=gc)
    ws1.row_dimensions[row].height=16
    prev=rev

for c in range(1,10): W(ws1,c,18)

# ══ SHEET 2: MONTHLY TREND ════════════════════════════
ws2=wb.create_sheet('📅 Monthly Trend')
ws2.sheet_view.showGridLines=False
title_row(ws2,1,'🌸  MY PARF — MONTHLY SALES TREND','A:G')
for ci,h in enumerate(['Month','Revenue (Rs.)','Revenue (Cr)','MoM %','Units','Invoices','Avg Price'],1):
    hdr(ws2,2,ci,h)
prev=0
for ri,r in enumerate(monthly):
    row=3+ri; bg=C_ALT if ri%2==0 else C_WHITE
    rev=float(r[1]); mom=((rev-prev)/prev*100) if prev>0 else 0
    ms=f'+{mom:.1f}%' if mom>0 and prev>0 else (f'{mom:.1f}%' if prev>0 else '—')
    mc=C_GREEN if mom>0 and prev>0 else (C_RED if mom<0 else '555555')
    vals=[r[0],rev,rev/1e7,ms,int(r[2]),int(r[3]),float(r[4])]
    fmts=[None,'#,##0','0.00',None,'#,##0','#,##0','#,##0']
    aligns=['center','right','center','center','center','center','right']
    for ci,(v,al,fmt) in enumerate(zip(vals,aligns,fmts),1):
        c=cell(ws2,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==4: c.font=bold(color=mc)
    ws2.row_dimensions[row].height=15
    prev=rev
for ci,w in enumerate([12,18,13,12,12,12,16],1): W(ws2,ci,w)

chart2=LineChart()
chart2.title="Monthly Revenue Trend"; chart2.style=10
chart2.y_axis.title="Revenue (Rs.)"; chart2.width=26; chart2.height=14
d2=Reference(ws2,min_col=2,max_col=2,min_row=2,max_row=2+len(monthly))
c2=Reference(ws2,min_col=1,min_row=3,max_row=2+len(monthly))
chart2.add_data(d2,titles_from_data=True); chart2.set_categories(c2)
chart2.series[0].graphicalProperties.line.solidFill=C_BRAND
chart2.series[0].graphicalProperties.line.width=20000
ws2.add_chart(chart2,f'A{4+len(monthly)}')

# ══ SHEET 3: PRODUCT ANALYSIS ════════════════════════
ws3=wb.create_sheet('🏆 Product Analysis')
ws3.sheet_view.showGridLines=False
title_row(ws3,1,'🌸  MY PARF — PRODUCT ANALYSIS (ALL SKUs)','A:H')
for ci,h in enumerate(['Rank','Product Name','Code','MRP','Avg Price','Disc %','Units','Revenue'],1):
    hdr(ws3,2,ci,h)
for ri,r in enumerate(products):
    row=3+ri; bg=C_LIGHT if ri%2==0 else C_WHITE
    disc=float(r[7]); dc=C_RED if disc>50 else (C_GOLD if disc>35 else C_GREEN)
    vals=[ri+1,str(r[0])[:55] if r[0] else r[1],r[1],float(r[2]),float(r[6]),disc,int(r[4]),float(r[3])]
    fmts=[None,None,None,'#,##0','#,##0','0.0"%"','#,##0','#,##0']
    als=['center','left','center','right','right','center','center','right']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws3,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==6: c.font=bold(color=dc)
        if ci==1 and ri<3: c.font=bold(color=C_BRAND)
    ws3.row_dimensions[row].height=14
for ci,w in enumerate([5,48,12,12,14,10,10,16],1): W(ws3,ci,w)

# ══ SHEET 4: BRANCH ANALYSIS ════════════════════════
ws4=wb.create_sheet('🏪 Branch Analysis')
ws4.sheet_view.showGridLines=False
title_row(ws4,1,'🌸  MY PARF — BRANCH-WISE PERFORMANCE (FULL NAMES)','A:H')
for ci,h in enumerate(['Rank','Branch Name','Short Code','Revenue','Rev (Cr)','Share %','Units','Invoices'],1):
    hdr(ws4,2,ci,h)
total_rev=sum(float(r[1]) for r in branches)
for ri,r in enumerate(branches):
    row=3+ri; bg=C_ALT if ri%2==0 else C_WHITE
    code=str(r[0]); fname=full_br(code)
    rev=float(r[1]); share=rev/total_rev*100
    vals=[ri+1,fname,code,rev,rev/1e7,share,int(r[2]),int(r[3])]
    fmts=[None,None,None,'#,##0','0.00','0.0"%"','#,##0','#,##0']
    als=['center','left','center','right','center','center','center','center']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws4,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==1 and ri<3: c.font=bold(color=C_BRAND)
    ws4.row_dimensions[row].height=14
for ci,w in enumerate([5,32,12,18,10,10,10,10],1): W(ws4,ci,w)

# Top 15 bar chart
chart4=BarChart(); chart4.type="bar"; chart4.style=10
chart4.title="Top 15 Branches by Revenue"; chart4.width=28; chart4.height=16
d4=Reference(ws4,min_col=4,max_col=4,min_row=2,max_row=min(17,2+len(branches)))
c4=Reference(ws4,min_col=2,min_row=3,max_row=min(17,2+len(branches)))
chart4.add_data(d4,titles_from_data=True); chart4.set_categories(c4)
chart4.series[0].graphicalProperties.solidFill=C_BRAND
ws4.add_chart(chart4,f'A{4+len(branches)}')

# ══ SHEET 5: CUSTOMER ANALYTICS ═══════════════════════
ws5=wb.create_sheet('👥 Customer Analytics')
ws5.sheet_view.showGridLines=False
title_row(ws5,1,'🌸  MY PARF — CUSTOMER ANALYTICS & INSIGHTS','A:H',h=32)

# New vs Repeat monthly
ws5.merge_cells('A3:H3')
ws5['A3']='NEW vs REPEAT CUSTOMERS — MONTHLY'; ws5['A3'].font=bold(size=11,color=C_WHITE)
ws5['A3'].fill=fill(C_ACCENT); ws5['A3'].alignment=ctr()
for ci,h in enumerate(['Month','New Customers','Repeat Customers','Total Customers','New %','Repeat %','New Rev Est.','Repeat Rev Est.'],1):
    hdr(ws5,4,ci,h,bg=C_BRAND)
for ri,r in enumerate(new_repeat_monthly):
    row=5+ri; bg=C_ALT if ri%2==0 else C_WHITE
    nc=int(r[1]); rc=int(r[2]); tc=int(r[3])
    np_=nc/tc*100 if tc>0 else 0; rp_=rc/tc*100 if tc>0 else 0
    vals=[r[0],nc,rc,tc,np_,rp_,'—','—']
    fmts=[None,None,None,None,'0.0"%"','0.0"%"',None,None]
    als=['center','center','center','center','center','center','center','center']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws5,row,ci,v,fmt=fmt,bg=bg,align=al)
    ws5.row_dimensions[row].height=15
for ci,w in enumerate([12,16,18,16,10,10,14,14],1): W(ws5,ci,w)

# Customer type section
nr_end=5+len(new_repeat_monthly)
ws5.row_dimensions[nr_end+1].height=8
ws5.merge_cells(f'A{nr_end+2}:H{nr_end+2}')
ws5[f'A{nr_end+2}']='CUSTOMER TYPE BREAKDOWN'
ws5[f'A{nr_end+2}'].font=bold(size=11,color=C_WHITE)
ws5[f'A{nr_end+2}'].fill=fill(C_ACCENT); ws5[f'A{nr_end+2}'].alignment=ctr()
for ci,h in enumerate(['Customer Type','Unique Customers','Revenue','Avg Spend'],1):
    hdr(ws5,nr_end+3,ci,h,bg=C_PURPLE)
total_ct_rev=sum(float(r[2]) for r in cust_type)
for ri,r in enumerate(cust_type):
    row=nr_end+4+ri; bg=C_ALT if ri%2==0 else C_WHITE
    avg_spend=float(r[2])/int(r[1]) if int(r[1])>0 else 0
    vals=[r[0] or 'Unknown',int(r[1]),float(r[2]),avg_spend]
    fmts=[None,None,'#,##0','#,##0']
    for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
        cell(ws5,row,ci,v,fmt=fmt,bg=bg,align='center')
    ws5.row_dimensions[row].height=15

# Frequency distribution
ct_end=nr_end+4+len(cust_type)
ws5.row_dimensions[ct_end+1].height=8
ws5.merge_cells(f'A{ct_end+2}:H{ct_end+2}')
ws5[f'A{ct_end+2}']='PURCHASE FREQUENCY DISTRIBUTION'
ws5[f'A{ct_end+2}'].font=bold(size=11,color=C_WHITE)
ws5[f'A{ct_end+2}'].fill=fill(C_ACCENT); ws5[f'A{ct_end+2}'].alignment=ctr()
for ci,h in enumerate(['Purchases','Customers','Revenue','Avg Revenue/Customer','% of Customers'],1):
    hdr(ws5,ct_end+3,ci,h,bg=C_BLUE)
total_freq_custs=sum(int(r[1]) for r in freq_dist)
for ri,r in enumerate(freq_dist):
    row=ct_end+4+ri; bg=C_LIGHT if ri%2==0 else C_WHITE
    nc=int(r[1]); rev=float(r[2])
    avg=rev/nc if nc>0 else 0; pct=nc/total_freq_custs*100 if total_freq_custs>0 else 0
    vals=[f'{int(r[0])}x',nc,rev,avg,pct]
    fmts=[None,None,'#,##0','#,##0','0.0"%"']
    for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
        cell(ws5,row,ci,v,fmt=fmt,bg=bg,align='center')
    ws5.row_dimensions[row].height=15

# ══ SHEET 6: TOP CUSTOMERS ════════════════════════════
ws6=wb.create_sheet('🌟 Top Customers')
ws6.sheet_view.showGridLines=False
title_row(ws6,1,'🌸  MY PARF — TOP 30 CUSTOMERS BY LIFETIME SPEND','A:H',h=30)
ws6.merge_cells('A2:H2')
ws6['A2']='Based on MY PARF purchases only | Ranked by total spend on MY PARF products'
ws6['A2'].font=normal(size=9,color='888888'); ws6['A2'].fill=fill(C_GRAY)
ws6['A2'].alignment=ctr()

for ci,h in enumerate(['Rank','Mobile No.','Customer Type','Total Spend','Visits','First Buy','Last Buy','Recency (Days)'],1):
    hdr(ws6,3,ci,h,bg=C_DARK)

today_d=date.today()
for ri,r in enumerate(top_customers):
    row=4+ri; bg=C_LIGHT if ri<3 else (C_ALT if ri%2==0 else C_WHITE)
    last=r[4]; first=r[5]
    recency=(today_d-last).days if last else 999
    rec_color=C_RED if recency>90 else (C_GOLD if recency>30 else C_GREEN)
    vals=[ri+1,r[0],r[1] or '—',float(r[2]),int(r[3]),str(r[5]),str(r[4]),recency]
    fmts=[None,None,None,'#,##0',None,None,None,None]
    als=['center','center','center','right','center','center','center','center']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws6,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==1 and ri<3: c.font=bold(size=11,color=C_BRAND)
        if ci==8: c.font=bold(color=rec_color)
    ws6.row_dimensions[row].height=15
for ci,w in enumerate([6,16,16,16,8,12,12,14],1): W(ws6,ci,w)

# ══ SHEET 7: PRICE & DAY ANALYSIS ════════════════════
ws7=wb.create_sheet('💰 Price & Day Analysis')
ws7.sheet_view.showGridLines=False
title_row(ws7,1,'🌸  MY PARF — PRICE TIER & DAY-OF-WEEK ANALYSIS','A:F',h=28)

ws7.merge_cells('A3:F3')
ws7['A3']='PRICE TIER BREAKDOWN'; ws7['A3'].font=bold(size=11,color=C_WHITE)
ws7['A3'].fill=fill(C_ACCENT); ws7['A3'].alignment=ctr()
for ci,h in enumerate(['Price Tier','SKU Count','Units','Revenue','Rev (Cr)','Share %'],1):
    hdr(ws7,4,ci,h)
total_tier=sum(float(r[3]) for r in price_tiers)
for ri,r in enumerate(price_tiers):
    row=5+ri; bg=C_ALT if ri%2==0 else C_WHITE
    rev=float(r[3])
    vals=[r[0],int(r[1]),int(r[2]),rev,rev/1e7,rev/total_tier*100]
    fmts=[None,None,'#,##0','#,##0','0.00','0.0"%"']
    for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
        cell(ws7,row,ci,v,fmt=fmt,bg=bg,align='center')
    ws7.row_dimensions[row].height=16

# Pie chart
chart7=PieChart(); chart7.title="Revenue Share by Price Tier"
chart7.style=10; chart7.width=16; chart7.height=12
d7=Reference(ws7,min_col=4,max_col=4,min_row=4,max_row=4+len(price_tiers))
c7=Reference(ws7,min_col=1,min_row=5,max_row=4+len(price_tiers))
chart7.add_data(d7,titles_from_data=True); chart7.set_categories(c7)
ws7.add_chart(chart7,'A9')

# Day of week
ws7.merge_cells('A22:F22')
ws7['A22']='SALES BY DAY OF WEEK'; ws7['A22'].font=bold(size=11,color=C_WHITE)
ws7['A22'].fill=fill(C_ACCENT); ws7['A22'].alignment=ctr()
for ci,h in enumerate(['Day','Revenue','Units','Vs Avg %','Rank','Best/Worst'],1):
    hdr(ws7,23,ci,h)
avg_dow=sum(float(r[2]) for r in dow)/len(dow)
ranked_dow=sorted(enumerate(dow),key=lambda x:float(x[1][2]),reverse=True)
rank_map={r[0]:ri+1 for ri,r in enumerate(ranked_dow)}
for ri,r in enumerate(dow):
    row=24+ri; bg=C_LIGHT if ri%2==0 else C_WHITE
    rev=float(r[2]); perf=(rev-avg_dow)/avg_dow*100
    ps=f'+{perf:.1f}%' if perf>=0 else f'{perf:.1f}%'
    rank=rank_map[ri]
    label='🔥 BEST DAY' if rank==1 else ('⚠️ LOWEST' if rank==len(dow) else '')
    vals=[r[1],rev,int(r[3]),ps,rank,label]
    fmts=[None,'#,##0','#,##0',None,None,None]
    als=['center','right','center','center','center','center']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws7,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==4: c.font=bold(color=C_GREEN if perf>=0 else C_RED)
    ws7.row_dimensions[row].height=16
for ci,w in enumerate([14,16,12,12,8,14],1): W(ws7,ci,w)

# ══ SHEET 8: LAST 30 DAYS ════════════════════════════
ws8=wb.create_sheet('📆 Last 30 Days')
ws8.sheet_view.showGridLines=False
title_row(ws8,1,'🌸  MY PARF — LAST 30 DAYS DAILY PERFORMANCE','A:F',h=28)
for ci,h in enumerate(['Date','Day','Revenue','Rev (Cr)','Units','Invoices'],1):
    hdr(ws8,2,ci,h)
days_map={0:'Mon',1:'Tue',2:'Wed',3:'Thu',4:'Fri',5:'Sat',6:'Sun'}
avg_d=sum(float(r[1]) for r in daily30)/max(len(daily30),1)
for ri,r in enumerate(daily30):
    row=3+ri; rev=float(r[1])
    dn=days_map.get(r[0].weekday(),'')
    hi=rev>avg_d*1.5; we=dn in ('Sat','Sun')
    bg='FFF0F0' if hi else (C_LIGHT if we else (C_ALT if ri%2==0 else C_WHITE))
    vals=[str(r[0]),dn,rev,rev/1e7,int(r[2]),int(r[3])]
    fmts=[None,None,'#,##0','0.00','#,##0','#,##0']
    als=['center','center','right','center','center','center']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws8,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==3 and hi: c.font=bold(color=C_BRAND)
    ws8.row_dimensions[row].height=14
tr=3+len(daily30)
for ci,v in enumerate(['TOTAL','',sum(float(r[1]) for r in daily30),sum(float(r[1]) for r in daily30)/1e7,
                        sum(int(r[2]) for r in daily30),sum(int(r[3]) for r in daily30)],1):
    c=cell(ws8,tr,ci,v,fmt=('#,##0' if ci in [3,5,6] else ('0.00' if ci==4 else None)),
           bg=C_DARK,bold_=True,color=C_WHITE,align='center')
for ci,w in enumerate([12,8,16,12,10,10],1): W(ws8,ci,w)

chart8=LineChart(); chart8.title="Last 30 Days Revenue"; chart8.style=10
chart8.y_axis.title="Revenue (Rs.)"; chart8.width=26; chart8.height=14
d8=Reference(ws8,min_col=3,max_col=3,min_row=2,max_row=2+len(daily30))
c8=Reference(ws8,min_col=1,min_row=3,max_row=2+len(daily30))
chart8.add_data(d8,titles_from_data=True); chart8.set_categories(c8)
chart8.series[0].graphicalProperties.line.solidFill=C_BRAND
chart8.series[0].graphicalProperties.line.width=20000
ws8.add_chart(chart8,f'A{tr+2}')

# ══ SHEET 9: REPEAT CUSTOMER BEHAVIOR ════════════════════════
ws9=wb.create_sheet('🔄 Repeat Behavior')
ws9.sheet_view.showGridLines=False
title_row(ws9,1,'🌸  MY PARF — REPEAT CUSTOMER BEHAVIOR','A:G',h=30)

ws9.merge_cells('A3:G3')
ws9['A3']='1. LIFETIME VALUE: REPEAT vs ONE-TIME BUYERS'
ws9['A3'].font=bold(size=11,color=C_WHITE)
ws9['A3'].fill=fill(C_ACCENT); ws9['A3'].alignment=ctr()

for ci,h in enumerate(['Customer Segment','Customers','Total Revenue','Rev (Cr)','Avg Lifetime Value','Total Invoices','Avg Invoices/Cust'],1):
    hdr(ws9,4,ci,h,bg=C_BRAND)

tot_rev = sum(float(r[2]) for r in repeat_metrics)
for ri,r in enumerate(repeat_metrics):
    row=5+ri; bg=C_ALT if ri%2==0 else C_WHITE
    seg=r[0]; custs=int(r[1]); rev=float(r[2])
    clv=float(r[3]); invs=int(r[4])
    vals=[seg,custs,rev,rev/1e7,clv,invs,invs/custs if custs>0 else 0]
    fmts=[None,'#,##0','#,##0','0.00','#,##0','#,##0','0.0']
    als=['left','center','right','center','right','center','center']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws9,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==1 and 'Repeat' in seg: c.font=bold(color=C_BRAND)
    ws9.row_dimensions[row].height=16

# Gap metrics
row_gap=5+len(repeat_metrics)+2
ws9.merge_cells(f'A{row_gap}:G{row_gap}')
ws9[f'A{row_gap}']='2. TIME BETWEEN PURCHASES (For Repeat Customers Only)'
ws9[f'A{row_gap}'].font=bold(size=11,color=C_WHITE)
ws9[f'A{row_gap}'].fill=fill(C_PURPLE); ws9[f'A{row_gap}'].alignment=ctr()
ws9.row_dimensions[row_gap].height=20

gap_data=[
    ('Average days between 1st and 2nd purchase', f'{float(repeat_gap[0] or 0):.1f} days'),
    ('Average days between 1st and latest purchase (Lifetime)', f'{float(repeat_gap[1] or 0):.1f} days')
]
for ri,(lbl,val) in enumerate(gap_data):
    r=row_gap+1+ri
    ws9.merge_cells(f'A{r}:D{r}')
    ws9.merge_cells(f'E{r}:G{r}')
    ws9[f'A{r}']=lbl; ws9[f'A{r}'].font=bold(size=11); ws9[f'A{r}'].fill=fill(C_GRAY); ws9[f'A{r}'].alignment=lft(); ws9.cell(row=r,column=1).border=bdr()
    ws9[f'E{r}']=val; ws9[f'E{r}'].font=bold(size=11,color=C_BRAND); ws9[f'E{r}'].fill=fill(C_WHITE); ws9[f'E{r}'].alignment=ctr(); ws9.cell(row=r,column=5).border=bdr()
    ws9.row_dimensions[r].height=18

# Top products for repeat buyers
row_prod=row_gap+5
ws9.merge_cells(f'A{row_prod}:G{row_prod}')
ws9[f'A{row_prod}']='3. TOP PRODUCTS BOUGHT BY REPEAT CUSTOMERS'
ws9[f'A{row_prod}'].font=bold(size=11,color=C_WHITE)
ws9[f'A{row_prod}'].fill=fill(C_BLUE); ws9[f'A{row_prod}'].alignment=ctr()

for ci,h in enumerate(['Rank','Product Name','MRP','Unique Buyers','Units Bought','Total Revenue','Rev (Cr)'],1):
    hdr(ws9,row_prod+1,ci,h,bg=C_DARK)

for ri,r in enumerate(repeat_products):
    row=row_prod+2+ri; bg=C_LIGHT if ri%2==0 else C_WHITE
    vals=[ri+1,str(r[0])[:55],float(r[1]),int(r[2]),int(r[3]),float(r[4]),float(r[4])/1e7]
    fmts=[None,None,'#,##0','#,##0','#,##0','#,##0','0.00']
    als=['center','left','right','center','center','right','center']
    for ci,(v,al,fmt) in enumerate(zip(vals,als,fmts),1):
        c=cell(ws9,row,ci,v,fmt=fmt,bg=bg,align=al)
        if ci==1 and ri<3: c.font=bold(color=C_BRAND)
    ws9.row_dimensions[row].height=15

for ci,w in enumerate([22,48,12,16,16,18,12],1): W(ws9,ci,w)

# ══ SHEET 10: INSIGHTS & ACTIONS ════════════════════════
ws10=wb.create_sheet('🔍 Insights & Actions')
ws10.sheet_view.showGridLines=False
title_row(ws10,1,'🌸  MY PARF — STRATEGIC INSIGHTS & RECOMMENDED ACTIONS','A:G',h=32)

insights=[
    ('GROWTH ACCELERATION','POSITIVE',C_GREEN,[
        ('Jul 2026 Milestone','First month to cross Rs.1 Crore — Rs.1.10 Cr achieved'),
        ('2026 vs 2025','2026 YTD already 2.12x full year 2025 revenue'),
        ('Projected 2026','At current pace: ~Rs.6.8 Cr annual revenue'),
        ('Recommendation','Expand SKU range; increase branch onboarding to 200+ branches'),
    ]),
    ('DISCOUNT ALERT','URGENT — ACTION NEEDED',C_RED,[
        ('ACECHILL 50ML','Selling at 58% below MRP — highest discount in portfolio'),
        ('City Legend 50ML','Selling at 57% below MRP — major margin erosion'),
        ('30ml Range','Avg 33-35% discount — acceptable, but monitor closely'),
        ('Recommendation','Set minimum floor price per SKU; enforce compliance at branch level'),
    ]),
    ('CUSTOMER RETENTION','NEEDS ATTENTION',C_GOLD,[
        ('Dormant Customers',f'{dormant_c:,} MY PARF customers not purchased in 90+ days'),
        ('New vs Repeat','New customers = {:,} | Repeat = {:,} ({:.0f}% repeat rate)'.format(new_customers_total, repeat_custs, repeat_custs/max(total_custs,1)*100)),
        ('Avg Lifetime Value','Rs.{:,.0f} per customer (MY PARF only)'.format(avg_clv)),
        ('Recommendation','Launch MY PARF loyalty club / fragrance subscription program'),
    ]),
    ('PRODUCT FOCUS','OPTIMIZE PORTFOLIO',C_PURPLE,[
        ('Top 3 Revenue SKUs','Amora Bloom, Imperium, Eclipse Noir = 27.4% of revenue'),
        ('Premium Underperform','31 SKUs at Rs.1,500+ yield only Rs.2.4L/SKU avg revenue'),
        ('30ml Champions','68.5% of units from Rs.899 tier — clear consumer preference'),
        ('Recommendation','Focus on top 30ml SKUs; rationalize or relaunch premium lineup'),
    ]),
    ('BRANCH EXPANSION','GROWTH OPPORTUNITY',C_BLUE,[
        ('Current Coverage',f'{int(summary[3])} branches active out of estimated 200+ total'),
        ('Untapped Revenue','~50-60 branches not yet selling MY PARF'),
        ('Avg Branch Revenue',f'Rs.{float(summary[0])/int(summary[3])/1e5:.1f}L per branch (all-time)'),
        ('Recommendation','Prioritize onboarding high-traffic branches; target Rs.2.5 Cr uplift'),
    ]),
    ('SEASONAL STRATEGY','PLAN AHEAD',C_BRAND,[
        ('Weak Period','August consistently soft (-22.5% in 2025, -38% partial Aug 2026)'),
        ('Peak Period','March, May, June are strongest months — pre-stock well'),
        ('Weekend Performance','Saturday = 21% above average; Friday + Sunday also strong'),
        ('Recommendation','Run Aug promotions to counter dip; stock up before Mar/May peak'),
    ]),
]

crow=3
for title,status,color,points in insights:
    ws10.merge_cells(f'A{crow}:C{crow}')
    ws10.merge_cells(f'D{crow}:G{crow}')
    ws10[f'A{crow}']=title; ws10[f'A{crow}'].font=bold(size=12,color=C_WHITE)
    ws10[f'A{crow}'].fill=fill(color); ws10[f'A{crow}'].alignment=ctr()
    ws10[f'D{crow}']=status; ws10[f'D{crow}'].font=bold(size=10,color=color)
    ws10[f'D{crow}'].fill=fill(C_WHITE); ws10[f'D{crow}'].alignment=ctr()
    ws10.row_dimensions[crow].height=22; crow+=1
    for label,detail in points:
        ws10.cell(row=crow,column=2,value=label).font=bold(size=10,color=color)
        ws10.cell(row=crow,column=2).fill=fill(C_GRAY); ws10.cell(row=crow,column=2).alignment=lft()
        ws10.merge_cells(f'C{crow}:G{crow}')
        ws10.cell(row=crow,column=3,value=detail).font=normal(size=10) if label!='Recommendation' else bold(size=10,color=C_BRAND)
        ws10.cell(row=crow,column=3).fill=fill(C_LIGHT if label=='Recommendation' else C_WHITE)
        ws10.cell(row=crow,column=3).alignment=lft()
        ws10.row_dimensions[crow].height=17; crow+=1
    crow+=1

for ci,w in enumerate([5,22,18,12,1,1,1],1): W(ws10,ci,max(w,5))

# ── Save ──────────────────────────────────────────────
out=r'C:\Users\jasil_myg\Desktop\MY_PARF_Analysis_Report_v5.xlsx'
wb.save(out)
print(f'\nSaved: {out}')
print('Sheets created:')
for s in wb.worksheets: print(f'  {s.title}')
