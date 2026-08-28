from django.http import JsonResponse
from django.core.cache import cache
import pandas as pd

# ── ClickHouse helper ─────────────────────────────────────────────────────────
def _get_ch():
    from analytics.clickhouse_service import get_ch_client
    return get_ch_client()


# ── Product → Display-category mapping (same as before) ──────────────────────
PRODUCT_CAT_MAPPING = {
    'AC OUTDOOR': 'OTHER', 'ACC BGN': 'ACC', 'ACC ZRD': 'ACC', 'AIR CONDITIONER': 'CE',
    'BT SPEAKERS': 'ACC', 'CAMERA': 'OTHER', 'CROCKERY': 'CE', 'DEMO': 'DEMO',
    'DESKTOP': 'OTHER', 'DRIER-DW-FREEZER': 'CE', 'FIXED ASSETS': 'OTHER', 'FOC': 'OTHER',
    'GDP-SPARE': 'OTHER', 'GIFT ITEMS': 'OTHER', 'HOME APPLIANCES': 'CE', 'HOME THEATRE': 'ACC',
    'HVA': 'ACC', 'IT PRODUCT': 'OTHER', 'Items': 'OTHER', 'LAPTOP': 'LAPTOP',
    'LAPTOP BAG': 'OTHER', 'LENS & BODY': 'OTHER', 'MICROWAVE OVEN': 'CE', 'MOBILE': 'MOBILE',
    'MONITOR': 'OTHER', 'MYG DOMO': 'RIG', 'PRINTER': 'RIG', 'PROJECTOR': 'RIG',
    'PROTECTION': 'OTHER', 'RECHARGE': 'OTHER', 'REFRIGERATORS': 'CE', 'RIG': 'RIG',
    'SCHEME': 'OTHER', 'SERVICE': 'OTHER', 'SIM': 'OTHER', 'SIM INVENTORY': 'OTHER',
    'SMALL APPLIANCES': 'CE', 'SMART WATCH': 'ACC', 'SPARE': 'SPARE', 'STANDBY': 'OTHER',
    'STATIONERY ITEMS': 'OTHER', 'TABLET': 'TABLET', 'TOTAL LOSS': 'OTHER', 'TV': 'CE',
    'WASHING MACHINES': 'CE', 'WATCH': 'OTHER', 'P&G': 'ACC', 'IT ACCESSORIES': 'ACC',
    'CARE PLUS': 'VALUE ADDED SERVICE', 'SERVICE CHARGES': 'OTHER', 'DTH': 'OTHER',
    'STABILIZER': 'CE', 'SCRAP': 'OTHER', 'D SPARE': 'SPARE', 'C SPARE': 'SPARE',
    'MYG VERSE': 'OTHER', 'DEMO LAPTOP': 'DEMO', 'EARBUDS': 'ACC',
    'RECHARGE INVENTORY': 'OTHER', 'CEGI': 'OTHER', 'STELLR VOUCHERS': 'RIG',
    'DUMMY': 'OTHER', 'DISH WASHER': 'CE', 'DRYER': 'CE', 'FREEZER': 'CE',
    'WET GRINDER': 'CE', 'DEMO ACCESSORIES': 'DEMO',
    'OSG WARRANTY': 'VALUE ADDED SERVICE', 'LG AMC': 'VALUE ADDED SERVICE',
    'PROTECT MAX': 'VALUE ADDED SERVICE', 'PERFUME': 'ACC', 'CONTRACT WORK': 'OTHER',
    'DEMO HA': 'DEMO', 'DIY': 'DEMO', 'MOBILE ANTIVIRUS': 'OTHER', 'HA ACCESSORIES': 'CE',
    'HOUSE HOLD': 'CE', 'TOTAL SECURITY': 'OTHER', 'KASPERSKY': 'OTHER',
    'QUICKHEAL': 'OTHER', 'CCTV': 'OTHER', 'DETERGENT': 'OTHER', 'ACCESSORIES': 'ACC',
    'AUDIO': 'ACC', 'EAR WEARABLES': 'ACC', 'FRAGRANCE': 'ACC', 'GAMING': 'ACC',
    'GLAMSHIELD': 'ACC', 'OFFER KIT': 'ACC', 'PERSONAL CARE': 'ACC',
    'PRINTER ACCESSORIES': 'ACC', 'STORAGE DEVICES': 'ACC',
}
EXCLUDE_CATS = {'SPARE', 'DEMO', 'OTHERS'}


# ── Date-range helpers ────────────────────────────────────────────────────────
def _get_months(q):
    return {'JFM': [1,2,3], 'AMJ': [4,5,6], 'JAS': [7,8,9], 'OND': [10,11,12]}.get(q, [1,2,3])

def _date_filter_sql(comp_type, val, year=2026, alias='s'):
    """Return a SQL WHERE fragment for date filtering on azure_sales_report."""
    col = f"toDate({alias}.date)"
    if comp_type == 'monthly':
        return f"toMonth({col}) = {int(val)} AND toYear({col}) = {year}"
    elif comp_type == 'quarterly':
        months = _get_months(val)
        m_list = ','.join(str(m) for m in months)
        return f"toMonth({col}) IN ({m_list}) AND toYear({col}) = {year}"
    elif comp_type == 'yearly':
        return f"toYear({col}) = {int(val)}"
    elif comp_type == 'fy':
        return f"{col} >= toDate('{int(val)}-04-01') AND {col} <= toDate('{int(val)+1}-03-31')"
    elif comp_type == 'custom':
        if '|' in val:
            start_date, end_date = val.split('|', 1)
            return f"{col} >= toDate('{start_date}') AND {col} <= toDate('{end_date}')"
        return "1=1"
    return "1=1"


# ── Main ClickHouse fetch ─────────────────────────────────────────────────────
def _fetch_period(ch, comp_type, val, year, brand_filter='', branch_filter='', category_filter='', product_filter=''):
    """
    Fetch from ClickHouse. Filters can be comma-separated multi-values.
    Note: category_filter is applied AFTER the PRODUCT_CAT_MAPPING in Python,
    since category is derived from product names, not the DB category column.
    """
    date_cond = _date_filter_sql(comp_type, val, year)
    
    def _in(col, val):
        vals = [v.strip() for v in val.split(',') if v.strip()] if val else []
        if not vals: return None
        escaped = "','".join(v.replace("'", "''") for v in vals)
        return f"{col} IN ('{escaped}')"
    
    extra = [c for c in [
        _in('m.brand',   brand_filter),
        _in('s.branch',  branch_filter),
        # Note: category_filter applied post-fetch in Python
        _in('m.product', product_filter),
    ] if c]

    where = f"toDate(s.date) != toDate('1970-01-01') AND {date_cond}"
    if extra:
        where += " AND " + " AND ".join(extra)

    sql = f"""
        SELECT
            if(isNull(m.product) OR m.product='', 'Unknown', m.product) AS product,
            if(isNull(m.brand) OR m.brand='', 'Unknown', m.brand)       AS brand,
            if(isNull(s.branch) OR s.branch='', 'Unknown', s.branch)    AS branch,
            toDate(s.date)                     AS sale_date,
            s.invoice_no                       AS invoice_no,
            toFloat64(s.qty)                   AS qty,
            toFloat64(s.sold_price)            AS sold_price
        FROM azure_sales_report s
        LEFT JOIN item_master m ON s.item_code = m.item_code
        WHERE {where}
    """
    rows = ch.query(sql).result_rows
    # columns: product, brand, branch, sale_date, invoice_no, qty, sold_price
    if not rows:
        return pd.DataFrame(columns=['category','product','brand','branch','sale_date','invoice_no','qty','sold_price']), 0

    df = pd.DataFrame(rows, columns=['product','brand','branch','sale_date','invoice_no','qty','sold_price'])
    df['qty']        = pd.to_numeric(df['qty'],        errors='coerce').fillna(0).astype(float)
    df['sold_price'] = pd.to_numeric(df['sold_price'], errors='coerce').fillna(0).astype(float)

    # Apply product → display-category mapping
    df['category'] = df['product'].map(PRODUCT_CAT_MAPPING).fillna('OTHER')

    inv_count = df['invoice_no'].nunique()
    return df, inv_count


def build_api_response(request):
    comp_type  = request.GET.get('type',     'monthly')
    base_val   = request.GET.get('base',     '4')
    comp_val   = request.GET.get('comp',     '5')
    base_year  = int(request.GET.get('base_year', 2026))
    comp_year  = int(request.GET.get('comp_year', 2026))
    cat        = request.GET.get('category', '')
    product    = request.GET.get('product',  '')
    brand      = request.GET.get('brand',    '')
    branch     = request.GET.get('branch',   '')
    rbm        = request.GET.get('rbm',      '')
    bdm        = request.GET.get('bdm',      '')

    cache_key = f"ent_ch_v2_{comp_type}_{base_val}_{comp_val}_{base_year}_{comp_year}_{cat}_{product}_{brand}_{branch}_{rbm}_{bdm}"
    cached = cache.get(cache_key)
    if cached:
        return cached

    ch = _get_ch()

    from .utils import get_branch_mappings
    code_to_name, name_to_code = get_branch_mappings(ch)

    # Resolve branch names → codes (from Branch filter)
    if branch:
        branch = ','.join([name_to_code.get(b.strip(), b.strip()) for b in branch.split(',')])

    # Resolve RBM filter → branch codes via branch_master
    if rbm:
        rbm_list = [r.strip() for r in rbm.split(',') if r.strip()]
        rbm_escaped = "','".join(r.replace("'", "''") for r in rbm_list)
        rbm_branches = ch.query(f"SELECT DISTINCT code FROM branch_master WHERE rbm IN ('{rbm_escaped}') AND code != ''").result_rows
        rbm_codes = [r[0] for r in rbm_branches]
        # Merge with existing branch filter
        if rbm_codes:
            existing = [b for b in branch.split(',') if b] if branch else []
            merged = list(set(existing + rbm_codes)) if existing else rbm_codes
            branch = ','.join(merged)

    # Resolve BDM filter → branch codes via branch_master
    if bdm:
        bdm_list = [b.strip() for b in bdm.split(',') if b.strip()]
        bdm_escaped = "','".join(b.replace("'", "''") for b in bdm_list)
        bdm_branches = ch.query(f"SELECT DISTINCT code FROM branch_master WHERE bdm IN ('{bdm_escaped}') AND code != ''").result_rows
        bdm_codes = [r[0] for r in bdm_branches]
        if bdm_codes:
            existing = [b for b in branch.split(',') if b] if branch else []
            merged = list(set(existing + bdm_codes)) if existing else bdm_codes
            branch = ','.join(merged)

    # ── Fetch base and comparison periods from ClickHouse ─────────────────────
    df_b, b_inv = _fetch_period(ch, comp_type, base_val, base_year, brand, branch, cat, product)
    df_c, c_inv = _fetch_period(ch, comp_type, comp_val, comp_year, brand, branch, cat, product)

    if not df_b.empty: df_b['branch'] = df_b['branch'].map(code_to_name).fillna(df_b['branch'])
    if not df_c.empty: df_c['branch'] = df_c['branch'].map(code_to_name).fillna(df_c['branch'])

    # Apply category filter AFTER mapping (mapped categories like MOBILE, CE, ACC etc.)
    if cat:
        cat_vals = [c.strip() for c in cat.split(',') if c.strip()]
        if cat_vals:
            df_b = df_b[df_b['category'].isin(cat_vals)]
            df_c = df_c[df_c['category'].isin(cat_vals)]

    # ── 1. KPIs ───────────────────────────────────────────────────────────────
    b_sales = float(df_b['sold_price'].sum())
    c_sales = float(df_c['sold_price'].sum())
    b_qty   = float(df_b['qty'].sum())
    c_qty   = float(df_c['qty'].sum())

    b_vas = float(df_b[df_b['category'] == 'VALUE ADDED SERVICE']['sold_price'].sum())
    c_vas = float(df_c[df_c['category'] == 'VALUE ADDED SERVICE']['sold_price'].sum())

    sales_growth = ((c_sales - b_sales) / b_sales * 100) if b_sales > 0 else 0
    qty_growth   = ((c_qty   - b_qty)   / b_qty   * 100) if b_qty   > 0 else 0
    vas_growth   = ((c_vas   - b_vas)   / b_vas   * 100) if b_vas   > 0 else 0

    b_asp      = b_sales / b_qty if b_qty > 0 else 0
    c_asp      = c_sales / c_qty if c_qty > 0 else 0
    asp_growth = ((c_asp - b_asp) / b_asp * 100) if b_asp > 0 else 0

    kpis = {
        "base":   {"sales": b_sales, "qty": b_qty, "vas": b_vas, "asp": b_asp, "inv": b_inv},
        "comp":   {"sales": c_sales, "qty": c_qty, "vas": c_vas, "asp": c_asp, "inv": c_inv},
        "growth": {"sales": sales_growth, "qty": qty_growth, "vas": vas_growth, "asp": asp_growth},
    }

    # ── 2. Trends ─────────────────────────────────────────────────────────────
    is_monthly = (comp_type == 'monthly')
    if is_monthly:
        if not df_b.empty: df_b['t'] = pd.to_datetime(df_b['sale_date']).dt.day
        if not df_c.empty: df_c['t'] = pd.to_datetime(df_c['sale_date']).dt.day
    else:
        if not df_b.empty: df_b['t'] = pd.to_datetime(df_b['sale_date']).dt.month
        if not df_c.empty: df_c['t'] = pd.to_datetime(df_c['sale_date']).dt.month

    trends = {"base": [], "comp": []}
    if not df_b.empty and 't' in df_b.columns:
        trends['base'] = (df_b.groupby('t')[['sold_price','qty']]
                           .sum().reset_index()
                           .rename(columns={'sold_price':'sales'})
                           .to_dict('records'))
    if not df_c.empty and 't' in df_c.columns:
        trends['comp'] = (df_c.groupby('t')[['sold_price','qty']]
                           .sum().reset_index()
                           .rename(columns={'sold_price':'sales'})
                           .to_dict('records'))

    # ── 3. Categories ─────────────────────────────────────────────────────────
    cat_base_dict = {}
    cats = []
    if not df_b.empty:
        cat_base_dict = df_b.groupby('category')['sold_price'].sum().to_dict()
    if not df_c.empty:
        cats = (df_c.groupby('category')[['sold_price','qty']]
                    .sum().reset_index()
                    .rename(columns={'sold_price':'sales'})
                    .sort_values('sales', ascending=False)
                    .to_dict('records'))

    # ── 4. Brands ─────────────────────────────────────────────────────────────
    brands = {}
    if not df_c.empty:
        brand_cats = df_c['category'].unique().tolist()
        for b_cat in brand_cats:
            b_df = df_c[df_c['category'] == b_cat]
            if not b_df.empty:
                brands[b_cat] = (b_df.groupby('brand')[['sold_price','qty']]
                                     .sum().reset_index()
                                     .rename(columns={'sold_price':'sales'})
                                     .sort_values('sales', ascending=False)
                                     .head(10)
                                     .to_dict('records'))

    # ── 5. Brand-level table ──────────────────────────────────────────────────
    tbl = []
    b_tbl = (df_b.groupby(['category','brand'])[['sold_price','qty']].sum().reset_index()
              .rename(columns={'sold_price':'b_rev','qty':'b_qty'})
             if not df_b.empty else pd.DataFrame(columns=['category','brand','b_rev','b_qty']))
    c_tbl = (df_c.groupby(['category','brand'])[['sold_price','qty']].sum().reset_index()
              .rename(columns={'sold_price':'c_rev','qty':'c_qty'})
             if not df_c.empty else pd.DataFrame(columns=['category','brand','c_rev','c_qty']))
    if not b_tbl.empty or not c_tbl.empty:
        merged = pd.merge(b_tbl, c_tbl, on=['category','brand'], how='outer').fillna(0)
        merged['v_rev'] = merged['c_rev'] - merged['b_rev']
        tbl = merged.to_dict('records')

    # ── 6. Category Scorecard ─────────────────────────────────────────────────
    scorecard = []
    b_cat_tbl = (df_b.groupby('category')[['sold_price','qty']].sum().reset_index()
                  .rename(columns={'sold_price':'b_rev','qty':'b_qty'})
                 if not df_b.empty else pd.DataFrame(columns=['category','b_rev','b_qty']))
    c_cat_tbl = (df_c.groupby('category')[['sold_price','qty']].sum().reset_index()
                  .rename(columns={'sold_price':'c_rev','qty':'c_qty'})
                 if not df_c.empty else pd.DataFrame(columns=['category','c_rev','c_qty']))

    if not b_cat_tbl.empty or not c_cat_tbl.empty:
        merged_cat = pd.merge(b_cat_tbl, c_cat_tbl, on='category', how='outer').fillna(0)
        
        # Enforce exact category list and order as requested by user
        exact_order = ['MOBILE', 'CE', 'LAPTOP', 'ACC', 'TABLET', 'OTHER', 'VALUE ADDED SERVICE', 'RIG']
        merged_cat = merged_cat[merged_cat['category'].isin(exact_order)]
        merged_cat['cat_order'] = pd.Categorical(merged_cat['category'], categories=exact_order, ordered=True)
        merged_cat = merged_cat.sort_values('cat_order').drop(columns=['cat_order'])

        merged_cat['rev_growth'] = merged_cat.apply(
            lambda r: ((r['c_rev']-r['b_rev'])/r['b_rev']*100) if r['b_rev'] > 0 else 0, axis=1)
        merged_cat['qty_growth'] = merged_cat.apply(
            lambda r: ((r['c_qty']-r['b_qty'])/r['b_qty']*100) if r['b_qty'] > 0 else 0, axis=1)
        total_c_rev = merged_cat['c_rev'].sum()
        merged_cat['share'] = merged_cat.apply(
            lambda r: (r['c_rev']/total_c_rev*100) if total_c_rev > 0 else 0, axis=1)
        merged_cat['b_asp'] = merged_cat.apply(
            lambda r: (r['b_rev']/r['b_qty']) if r['b_qty'] > 0 else 0, axis=1)
        merged_cat['c_asp'] = merged_cat.apply(
            lambda r: (r['c_rev']/r['c_qty']) if r['c_qty'] > 0 else 0, axis=1)

        if not merged_cat.empty:
            grand = {
                'category':    'Grand Total',
                'b_rev':       float(merged_cat['b_rev'].sum()),
                'c_rev':       float(total_c_rev),
                'b_qty':       float(merged_cat['b_qty'].sum()),
                'c_qty':       float(merged_cat['c_qty'].sum()),
            }
            grand['rev_growth'] = ((grand['c_rev']-grand['b_rev'])/grand['b_rev']*100) if grand['b_rev'] > 0 else 0
            grand['qty_growth'] = ((grand['c_qty']-grand['b_qty'])/grand['b_qty']*100) if grand['b_qty'] > 0 else 0
            grand['share']  = 100.0
            grand['b_asp']  = (grand['b_rev']/grand['b_qty']) if grand['b_qty'] > 0 else 0
            grand['c_asp']  = (grand['c_rev']/grand['c_qty']) if grand['c_qty'] > 0 else 0
            
            scorecard  = merged_cat.to_dict('records')
            scorecard.append(grand)

    # ── 7. Detailed product-brand-branch table (for Custom Report and Excel) ──────
    b_dtbl = (df_b.groupby(['category','product','brand','branch'])[['sold_price','qty']].sum().reset_index()
               .rename(columns={'sold_price':'b_rev','qty':'b_qty'})
              if not df_b.empty else pd.DataFrame(columns=['category','product','brand','branch','b_rev','b_qty']))
    c_dtbl = (df_c.groupby(['category','product','brand','branch'])[['sold_price','qty']].sum().reset_index()
               .rename(columns={'sold_price':'c_rev','qty':'c_qty'})
              if not df_c.empty else pd.DataFrame(columns=['category','product','brand','branch','c_rev','c_qty']))
    detailed_table = (pd.merge(b_dtbl, c_dtbl, on=['category','product','brand','branch'], how='outer')
                        .fillna(0).to_dict('records'))

    final_response = {
        "kpis":                kpis,
        "trends":              trends,
        "categories":          cats,
        "categories_base_dict": cat_base_dict,
        "brands":              brands,
        "table":               tbl,
        "scorecard":           scorecard,
        "detailed_table":      detailed_table,
    }

    cache.set(cache_key, final_response, timeout=3600)
    return final_response


# ── Excel export (unchanged logic, just uses 'sold_price' key now) ────────────
import io
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN       = Side(style="thin",   color="000000")
THIN_BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FONT   = Font(name="Arial", bold=True,  size=9)
BODY_FONT  = Font(name="Arial", bold=False, size=9)
TOT_FONT   = Font(name="Arial", bold=True,  size=9)
NEG_FILL   = PatternFill("solid", fgColor="FFC7CE")
TOT_FILL   = PatternFill("solid", fgColor="BDD7EE")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
COL_WIDTHS = [24, 12, 12, 11, 11, 2, 12, 12, 11, 11]

def _c(ws, row, col, value=None, font=None, fill=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if fmt:    cell.number_format = fmt
    return cell

def write_header(ws, row, label_name, b_val, c_val, sc=1):
    def hdr(r, c, v, colspan=1, rowspan=1):
        if colspan > 1 or rowspan > 1:
            ws.merge_cells(start_row=r, start_column=c, end_row=r+rowspan-1, end_column=c+colspan-1)
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = HDR_FONT; cell.fill = WHITE_FILL
        cell.border = THIN_BDR; cell.alignment = CENTER
        return cell
    hdr(row,   sc,   label_name, rowspan=2)
    hdr(row,   sc+1, b_val); hdr(row, sc+2, c_val)
    hdr(row,   sc+3, "GROWTH %", rowspan=2)
    hdr(row,   sc+4, f"SHARE {c_val}")
    ws.cell(row=row,   column=sc+5).border = Border()
    ws.cell(row=row+1, column=sc+5).border = Border()
    hdr(row,   sc+6, b_val); hdr(row, sc+7, c_val)
    hdr(row,   sc+8, "GROWTH %", rowspan=2)
    hdr(row,   sc+9, f"SHARE {c_val}")
    hdr(row+1, sc+1, "AMT"); hdr(row+1, sc+2, "AMT")
    hdr(row+1, sc+4, "AMT"); hdr(row+1, sc+6, "QTY")
    hdr(row+1, sc+7, "QTY"); hdr(row+1, sc+9, "QTY")
    return row + 2

def write_rows(ws, df, label_col, row, sc=1):
    for _, r in df.iterrows():
        is_tot = str(r[label_col]).strip().lower() == "grand total"
        fill  = TOT_FILL  if is_tot else WHITE_FILL
        font  = TOT_FONT  if is_tot else BODY_FONT
        def put(col, val, fmt=None, growth=False):
            v = None if (isinstance(val, float) and pd.isna(val)) else val
            cf = _c(ws, row, col, v, font=font, fill=fill,
                    align=RIGHT if col > sc else LEFT, border=THIN_BDR, fmt=fmt)
            if growth and v is not None and not is_tot:
                cf.fill = NEG_FILL if v < 0 else WHITE_FILL
            return cf
        put(sc,   r[label_col])
        put(sc+1, r["APR_AMT"], '#,##0.00')
        put(sc+2, r["MAY_AMT"], '#,##0.00')
        put(sc+3, r["AMT_GRW"], '0.00"%"', growth=True)
        put(sc+4, r["AMT_SHR"], '0.00"%"')
        ws.cell(row=row, column=sc+5)
        put(sc+6, r["APR_QTY"], '#,##0')
        put(sc+7, r["MAY_QTY"], '#,##0')
        put(sc+8, r["QTY_GRW"], '0.00"%"', growth=True)
        put(sc+9, r["QTY_SHR"], '0.00"%"')
        row += 1
    return row

def write_table(ws, df, label_col, section_label, b_val, c_val, start_row=1, sc=1):
    row = write_header(ws, start_row, section_label, b_val, c_val, sc)
    row = write_rows(ws, df, label_col, row, sc)
    for i, w in enumerate(COL_WIDTHS, start=sc):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=start_row+2, column=sc+1)
    return row + 1


def generate_dashboard_excel(request):
    from openpyxl import Workbook
    data      = build_api_response(request)
    comp_type = request.GET.get('type', 'monthly')
    base_val  = request.GET.get('base', '4')
    comp_val  = request.GET.get('comp', '5')
    base_year = request.GET.get('base_year', '2026')
    comp_year = request.GET.get('comp_year', '2026')
    
    # Prepend year if not yearly/fy
    b_label = base_val if comp_type in ('yearly', 'fy') else f"{base_val} {base_year}"
    c_label = comp_val if comp_type in ('yearly', 'fy') else f"{comp_val} {comp_year}"

    scorecard = data.get('scorecard', [])
    det_table = data.get('detailed_table', [])

    def make_sheet_df(data_rows, group_col, label_col_name):
        grouped = {}
        total_c = sum(r.get('c_rev', 0) for r in data_rows)
        total_cq = sum(r.get('c_qty', 0) for r in data_rows)
        
        for r in data_rows:
            g = r.get(group_col)
            if not g: continue
            if g not in grouped:
                grouped[g] = {'b_rev':0, 'c_rev':0, 'b_qty':0, 'c_qty':0}
            grouped[g]['b_rev'] += r.get('b_rev', 0)
            grouped[g]['c_rev'] += r.get('c_rev', 0)
            grouped[g]['b_qty'] += r.get('b_qty', 0)
            grouped[g]['c_qty'] += r.get('c_qty', 0)
            
        out = []
        for g, vals in grouped.items():
            b_rev = vals['b_rev']; c_rev = vals['c_rev']
            b_qty = vals['b_qty']; c_qty = vals['c_qty']
            out.append({
                label_col_name: g,
                'APR_AMT': round(b_rev/1e7, 2),
                'MAY_AMT': round(c_rev/1e7, 2),
                'AMT_GRW': round(((c_rev-b_rev)/b_rev*100) if b_rev>0 else 0, 2),
                'AMT_SHR': round((c_rev/total_c*100) if total_c>0 else 0, 2),
                'APR_QTY': int(b_qty),
                'MAY_QTY': int(c_qty),
                'QTY_GRW': round(((c_qty-b_qty)/b_qty*100) if b_qty>0 else 0, 2),
                'QTY_SHR': round((c_qty/total_cq*100) if total_cq>0 else 0, 2)
            })
        out.sort(key=lambda x: x['MAY_AMT'], reverse=True)
        
        tb_rev = sum(v['b_rev'] for v in grouped.values())
        tc_rev = sum(v['c_rev'] for v in grouped.values())
        tb_qty = sum(v['b_qty'] for v in grouped.values())
        tc_qty = sum(v['c_qty'] for v in grouped.values())
        
        if out:
            out.append({
                label_col_name: 'Grand Total',
                'APR_AMT': round(tb_rev/1e7, 2),
                'MAY_AMT': round(tc_rev/1e7, 2),
                'AMT_GRW': round(((tc_rev-tb_rev)/tb_rev*100) if tb_rev>0 else 0, 2),
                'AMT_SHR': 100.0,
                'APR_QTY': int(tb_qty),
                'MAY_QTY': int(tc_qty),
                'QTY_GRW': round(((tc_qty-tb_qty)/tb_qty*100) if tb_qty>0 else 0, 2),
                'QTY_SHR': 100.0
            })
        return pd.DataFrame(out)

    sheets_config = [
        ("1_Category Summary",  det_table,                                   'category', "PDT CAT"),
        ("2_Mobile Brand",      [r for r in det_table if r['category']=='MOBILE'], 'brand',    "MOBILE"),
        ("3_Laptop Brand",      [r for r in det_table if r['category']=='LAPTOP'], 'brand',    "LAPTOP"),
        ("4_ACC Brand",         [r for r in det_table if r['category']=='ACC'],    'brand',    "ACC"),
        ("5_ACC Category",      [r for r in det_table if r['category']=='ACC'],    'product',  "ACC"),
        ("6_CE Category",       [r for r in det_table if r['category']=='CE'],     'product',  "CE"),
        ("7_CE Brand",          [r for r in det_table if r['category']=='CE'],     'brand',    "CE"),
        ("8_TV Brand",          [r for r in det_table if 'TV' in r.get('product','').upper()], 'brand', "Tv"),
        ("9_Fridge Brand",      [r for r in det_table if 'REFRIGERATOR' in r.get('product','').upper()], 'brand', "Refrigerators"),
        ("10_WM Brand",         [r for r in det_table if 'WASHING' in r.get('product','').upper()], 'brand', "Washing Machines"),
        ("11_AC Brand",         [r for r in det_table if 'AIR CONDITIONER' in r.get('product','').upper() or 'AC ' in r.get('product','').upper()], 'brand', "Air Conditioner"),
        ("12_VAS Summary",      [r for r in det_table if r['category']=='VALUE ADDED SERVICE'], 'product', "VAS"),
    ]

    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    special_only = request.GET.get('special_only') == 'true'

    if not special_only:
        for sheet_name, rows, group_col, label_name in sheets_config:
            ws = wb.create_sheet(sheet_name)
            df_sheet = make_sheet_df(rows, group_col, label_name)
            if not df_sheet.empty:
                write_table(ws, df_sheet, label_name, label_name, b_label, c_label, start_row=1)
            else:
                ws.cell(row=1, column=1, value=f"No data for {sheet_name}")

    # ── Custom Detailed Report (Branch level) ──
    ws_det = wb.create_sheet("Special_Report" if special_only else "13_Detailed_Report")
    det_out = []
    for r in det_table:
        b_rev = r.get('b_rev',0)
        c_rev = r.get('c_rev',0)
        b_qty = r.get('b_qty',0)
        c_qty = r.get('c_qty',0)
        det_out.append({
            'Category': r.get('category',''),
            'Brand': r.get('brand',''),
            'Product': r.get('product',''),
            'Branch': r.get('branch',''),
            f'{b_label} AMT (Cr)': round(b_rev/1e7, 2),
            f'{c_label} AMT (Cr)': round(c_rev/1e7, 2),
            'AMT Growth %': round(((c_rev-b_rev)/b_rev*100) if b_rev>0 else 0, 2),
            f'{b_label} QTY': int(b_qty),
            f'{c_label} QTY': int(c_qty),
            'QTY Growth %': round(((c_qty-b_qty)/b_qty*100) if b_qty>0 else 0, 2)
        })
    if det_out:
        df_det = pd.DataFrame(det_out)
        df_det = df_det.sort_values(f'{c_label} AMT (Cr)', ascending=False)
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r_idx, row_data in enumerate(dataframe_to_rows(df_det, index=False, header=True), 1):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_det.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    
        # adjust column widths
        ws_det.column_dimensions['A'].width = 20
        ws_det.column_dimensions['B'].width = 20
        ws_det.column_dimensions['C'].width = 30
        ws_det.column_dimensions['D'].width = 30
    else:
        ws_det.cell(row=1, column=1, value="No detailed data")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Monthly_Analysis_{b_label}_vs_{c_label}.xlsx"
    return buf, filename
