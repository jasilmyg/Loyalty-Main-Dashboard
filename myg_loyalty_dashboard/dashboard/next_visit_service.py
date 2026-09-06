"""
next_visit_service.py
=====================
Smart Next-Visit Predictor — Service Layer

Uses 4 ClickHouse tables:
  - azure_invoice_report  : customer visit history (recency, frequency, monetary, gap)
  - azure_sales_report    : item-level purchases per customer
  - item_master           : item_code → category, brand, product
  - branch_master         : branch_code → branch_name, district, RBM, BDM

Prediction formula:
  predicted_start = last_visit + (avg_gap - gap_std)
  predicted_peak  = last_visit + avg_gap  (adjusted for preferred day-of-week)
  predicted_end   = last_visit + (avg_gap + gap_std)

Confidence bands:
  UPCOMING  — not yet due (days_since_visit < avg_gap - gap_std)
  HIGH      — due now    (within avg_gap ± gap_std)
  MEDIUM    — slightly overdue (avg_gap*1 < days_since < avg_gap*2)
  AT RISK   — long overdue   (days_since > avg_gap*2)
"""

import math
from datetime import date, timedelta
from typing import Optional

from analytics.clickhouse_service import get_ch_client

# Day-of-week names (ClickHouse toDayOfWeek: 1=Mon … 7=Sun)
DOW_NAMES = {1: "Monday", 2: "Tuesday", 3: "Wednesday",
             4: "Thursday", 5: "Friday", 6: "Saturday", 7: "Sunday"}

# Festival windows (approx Gregorian dates, refreshed annually)
FESTIVAL_WINDOWS_2026 = [
    ("Vishu",           date(2026, 4, 14), date(2026, 4, 16)),
    ("Eid ul-Fitr",     date(2026, 3, 30), date(2026, 4,  2)),
    ("Onam",            date(2026, 9,  5), date(2026, 9, 13)),
    ("Navratri",        date(2026, 9, 22), date(2026, 10, 1)),
    ("Diwali",          date(2026,10, 20), date(2026,10, 23)),
    ("Christmas",       date(2026,12, 24), date(2026,12, 26)),
    ("New Year",        date(2026,12, 31), date(2027,  1,  2)),
    ("Eid ul-Adha",     date(2026, 6,  6), date(2026,  6,  9)),
    ("Independence Day",date(2026, 8, 15), date(2026,  8, 16)),
    ("Republic Day",    date(2026, 1, 26), date(2026,  1, 27)),
]

VALID_MOBILE_FILTER = """
    length(customer_mobile) = 10
    AND customer_mobile NOT IN ('1313131313','0000000000','9999999999')
    AND customer_mobile != ''
"""


def _get_festival_in_window(start: date, end: date) -> Optional[str]:
    """Return the name of the nearest festival if one falls within the window."""
    for name, f_start, f_end in FESTIVAL_WINDOWS_2026:
        if start <= f_end and end >= f_start:
            return name
    return None


def _compute_confidence(days_since: int, avg_gap: float, gap_std: float) -> str:
    """Return confidence band label."""
    lower = max(1, avg_gap - gap_std)
    upper = avg_gap + gap_std
    if days_since < lower:
        return "UPCOMING"
    elif days_since <= upper:
        return "HIGH"
    elif days_since <= avg_gap * 2:
        return "MEDIUM"
    else:
        return "AT RISK"


def _confidence_color(confidence: str) -> str:
    return {
        "UPCOMING": "#3b82f6",
        "HIGH":     "#10b981",
        "MEDIUM":   "#f59e0b",
        "AT RISK":  "#ef4444",
    }.get(confidence, "#6b7280")


def _predicted_peak_date(last_visit: date, avg_gap: float, preferred_dow: int) -> date:
    """Shift avg_gap-based peak toward preferred day of week."""
    base = last_visit + timedelta(days=int(round(avg_gap)))
    # Search within ±4 days for the preferred day-of-week
    for delta in range(0, 5):
        for d in [base + timedelta(days=delta), base - timedelta(days=delta)]:
            if d.isoweekday() == preferred_dow:
                return d
    return base


def get_next_visit_predictions(
    branch: str = None,
    district: str = None,
    rbm: str = None,
    due_in_days: int = 7,
    limit: int = 500,
):
    """
    Main query: returns list of per-customer predictions, enriched with
    branch_master and item_master data.

    Filters:
      branch     — filter by last_branch code
      district   — filter by district (via branch_master)
      rbm        — filter by RBM name (via branch_master)
      due_in_days — only include customers whose predicted window starts
                    within the next N days
    """
    ch = get_ch_client()
    if not ch:
        return {"error": "ClickHouse unavailable", "rows": [], "kpis": {}}

    today = date.today()

    # ── Step 1: Resolve branch codes from district/rbm filters ───────────────
    branch_filter_codes = set()
    if district:
        dist_esc = district.replace("'", "''")
        rows = ch.query(
            f"SELECT DISTINCT code FROM branch_master WHERE district = '{dist_esc}' AND code != ''"
        ).result_rows
        branch_filter_codes.update(r[0] for r in rows)

    if rbm:
        rbm_esc = rbm.replace("'", "''")
        rows = ch.query(
            f"SELECT DISTINCT code FROM branch_master WHERE rbm = '{rbm_esc}' AND code != ''"
        ).result_rows
        branch_filter_codes.update(r[0] for r in rows)

    if branch:
        branch_filter_codes.add(branch)

    # Build branch WHERE clause
    if branch_filter_codes:
        codes_str = ",".join(f"'{c}'" for c in branch_filter_codes)
        branch_where = f"AND v.last_branch IN ({codes_str})"
    else:
        branch_where = ""

    # ── Step 2: Customer visit summary + gap stats ────────────────────────────
    summary_sql = f"""
    WITH
    visit_summary AS (
        SELECT
            customer_mobile,
            toDate(max(date))                       AS last_visit,
            today() - toDate(max(date))             AS days_since_visit,
            countDistinct(toDate(date))             AS visit_count,
            toInt64(sum(invoice_total))             AS total_spend,
            toInt64(avg(invoice_total))             AS avg_spend_per_visit,
            argMax(branch, date)                    AS last_branch
        FROM azure_invoice_report
        WHERE {VALID_MOBILE_FILTER}
          AND invoice_total > 0
          AND toDate(date) != toDate('1970-01-01')
        GROUP BY customer_mobile
        HAVING visit_count >= 2
    ),
    gap_stats AS (
        SELECT
            customer_mobile,
            round(avg(gap_days), 1)         AS avg_gap,
            round(stddevPop(gap_days), 1)   AS gap_std
        FROM (
            SELECT
                customer_mobile,
                dateDiff('day',
                    lagInFrame(toDate(date)) OVER (
                        PARTITION BY customer_mobile
                        ORDER BY toDate(date)
                    ),
                    toDate(date)
                ) AS gap_days
            FROM azure_invoice_report
            WHERE {VALID_MOBILE_FILTER}
              AND toDate(date) != toDate('1970-01-01')
        )
        WHERE gap_days IS NOT NULL AND gap_days > 0 AND gap_days < 730
        GROUP BY customer_mobile
    ),
    pref_dow AS (
        SELECT customer_mobile, dow
        FROM (
            SELECT
                customer_mobile,
                toDayOfWeek(toDate(date)) AS dow,
                count()                    AS cnt,
                row_number() OVER (PARTITION BY customer_mobile ORDER BY cnt DESC) AS rn
            FROM azure_invoice_report
            WHERE {VALID_MOBILE_FILTER}
              AND toDate(date) != toDate('1970-01-01')
            GROUP BY customer_mobile, dow
        )
        WHERE rn = 1
    )
    SELECT
        v.customer_mobile,
        v.last_visit,
        v.days_since_visit,
        v.visit_count,
        v.total_spend,
        v.avg_spend_per_visit,
        v.last_branch,
        coalesce(b.branch_name, v.last_branch) AS branch_name,
        coalesce(b.district, '')               AS district,
        coalesce(b.rbm, '')                    AS rbm,
        coalesce(b.bdm, '')                    AS bdm,
        coalesce(g.avg_gap, 45.0)              AS avg_gap,
        coalesce(g.gap_std,  10.0)             AS gap_std,
        coalesce(p.dow, 6)                     AS preferred_dow
    FROM visit_summary v
    LEFT JOIN branch_master b ON v.last_branch = b.code
    LEFT JOIN gap_stats     g ON v.customer_mobile = g.customer_mobile
    LEFT JOIN pref_dow      p ON v.customer_mobile = p.customer_mobile
    WHERE v.days_since_visit <= 730
    {branch_where}
    ORDER BY v.total_spend DESC
    LIMIT {limit * 3}
    """

    try:
        rows = ch.query(summary_sql).result_rows
    except Exception as e:
        return {"error": str(e), "rows": [], "kpis": {}}

    # ── Step 3: Fetch top category per customer (item_master join) ────────────
    # Batch: top 1000 customer mobiles to avoid enormous IN clause
    mobiles_sample = [r[0] for r in rows[:1000]]
    category_map = {}   # mobile → {"category": ..., "brand": ...}

    if mobiles_sample:
        mob_list = ",".join(f"'{m}'" for m in mobiles_sample)
        cat_sql = f"""
        SELECT customer_mobile, category, brand
        FROM (
            SELECT
                s.customer_mobile,
                coalesce(im.category, 'OTHER') AS category,
                coalesce(im.brand, '')          AS brand,
                count()                         AS cnt,
                row_number() OVER (
                    PARTITION BY s.customer_mobile
                    ORDER BY cnt DESC
                ) AS rn
            FROM azure_sales_report s
            LEFT JOIN item_master im ON s.item_code = im.item_code
            WHERE s.customer_mobile IN ({mob_list})
              AND im.category != ''
            GROUP BY s.customer_mobile, im.category, im.brand
        )
        WHERE rn = 1
        """
        try:
            cat_rows = ch.query(cat_sql).result_rows
            for cr in cat_rows:
                category_map[cr[0]] = {"category": cr[1], "brand": cr[2]}
        except Exception:
            pass  # Category enrichment is optional

    # ── Step 4: Build prediction objects ─────────────────────────────────────
    predictions = []
    for r in rows:
        mobile          = r[0]
        last_visit      = r[1]  # date object from ClickHouse
        days_since      = int(r[2])
        visit_count     = int(r[3])
        total_spend     = int(r[4])
        avg_spend       = int(r[5])
        last_branch     = r[6]
        branch_name     = r[7]
        district        = r[8]
        rbm             = r[9]
        bdm             = r[10]
        avg_gap         = float(r[11])
        gap_std         = float(r[12])
        preferred_dow   = int(r[13])

        # Compute prediction window
        window_start = last_visit + timedelta(days=max(1, int(avg_gap - gap_std)))
        window_end   = last_visit + timedelta(days=int(avg_gap + gap_std))
        peak_date    = _predicted_peak_date(last_visit, avg_gap, preferred_dow)

        # Days until peak from today
        days_to_peak = (peak_date - today).days

        # Apply due_in_days filter: show customers whose peak is within window
        # Include: past-due (negative) up to +due_in_days ahead
        if days_to_peak > due_in_days:
            continue

        confidence      = _compute_confidence(days_since, avg_gap, gap_std)
        conf_color      = _confidence_color(confidence)
        festival        = _get_festival_in_window(window_start, window_end)
        cat_info        = category_map.get(mobile, {})

        predictions.append({
            "mobile":           mobile,
            "mobile_masked":    f"XXXX-{mobile[-4:]}",
            "last_visit":       last_visit.strftime("%d %b %Y"),
            "days_since":       days_since,
            "visit_count":      visit_count,
            "total_spend":      total_spend,
            "avg_spend":        avg_spend,
            "last_branch":      last_branch,
            "branch_name":      branch_name,
            "district":         district,
            "rbm":              rbm,
            "bdm":              bdm,
            "avg_gap":          round(avg_gap, 1),
            "gap_std":          round(gap_std, 1),
            "preferred_dow":    DOW_NAMES.get(preferred_dow, "Saturday"),
            "window_start":     window_start.strftime("%d %b"),
            "window_end":       window_end.strftime("%d %b"),
            "peak_date":        peak_date.strftime("%d %b %Y"),
            "peak_date_short":  peak_date.strftime("%d %b"),
            "days_to_peak":     days_to_peak,
            "confidence":       confidence,
            "conf_color":       conf_color,
            "festival":         festival or "",
            "category":         cat_info.get("category", ""),
            "brand":            cat_info.get("brand", ""),
        })

        if len(predictions) >= limit:
            break

    # ── Step 5: Compute KPIs ──────────────────────────────────────────────────
    due_this_week  = sum(1 for p in predictions if 0 <= p["days_to_peak"] <= 7)
    high_conf      = sum(1 for p in predictions if p["confidence"] == "HIGH")
    medium_conf    = sum(1 for p in predictions if p["confidence"] == "MEDIUM")
    at_risk        = sum(1 for p in predictions if p["confidence"] == "AT RISK")
    upcoming       = sum(1 for p in predictions if p["confidence"] == "UPCOMING")
    avg_gap_all    = round(
        sum(p["avg_gap"] for p in predictions) / len(predictions), 1
    ) if predictions else 0

    kpis = {
        "total":          len(predictions),
        "due_this_week":  due_this_week,
        "high_conf":      high_conf,
        "medium_conf":    medium_conf,
        "at_risk":        at_risk,
        "upcoming":       upcoming,
        "avg_gap":        avg_gap_all,
        "data_source":    "ClickHouse",
    }

    return {"rows": predictions, "kpis": kpis, "error": None}


def get_single_customer_prediction(mobile: str) -> dict:
    """
    Full prediction card for a single customer mobile number.
    Used by the search API endpoint.
    Joins all 4 tables: azure_invoice_report, azure_sales_report,
    item_master, branch_master.
    """
    ch = get_ch_client()
    if not ch:
        return {"error": "ClickHouse unavailable"}

    today = date.today()
    m = mobile.strip()

    # ── Customer core stats ───────────────────────────────────────────────────
    core_sql = f"""
    SELECT
        customer_mobile,
        toDate(max(date))                       AS last_visit,
        today() - toDate(max(date))             AS days_since,
        countDistinct(toDate(date))             AS visit_count,
        toInt64(sum(invoice_total))             AS total_spend,
        toInt64(avg(invoice_total))             AS avg_spend,
        argMax(branch, date)                    AS last_branch
    FROM azure_invoice_report
    WHERE customer_mobile = '{m}'
      AND invoice_total > 0
      AND toDate(date) != toDate('1970-01-01')
    GROUP BY customer_mobile
    """
    try:
        core_rows = ch.query(core_sql).result_rows
    except Exception as e:
        return {"error": str(e)}

    if not core_rows:
        return {"error": f"No data found for mobile {m}"}

    r = core_rows[0]
    last_visit   = r[1]
    days_since   = int(r[2])
    visit_count  = int(r[3])
    total_spend  = int(r[4])
    avg_spend    = int(r[5])
    last_branch  = r[6]

    # ── Gap stats ─────────────────────────────────────────────────────────────
    gap_sql = f"""
    SELECT round(avg(gap_days), 1), round(stddevPop(gap_days), 1)
    FROM (
        SELECT
            dateDiff('day',
                lagInFrame(toDate(date)) OVER (ORDER BY toDate(date)),
                toDate(date)
            ) AS gap_days
        FROM azure_invoice_report
        WHERE customer_mobile = '{m}'
          AND toDate(date) != toDate('1970-01-01')
    )
    WHERE gap_days IS NOT NULL AND gap_days > 0 AND gap_days < 730
    """
    try:
        gap_row = ch.query(gap_sql).result_rows
        avg_gap  = float(gap_row[0][0] or 45.0) if gap_row else 45.0
        gap_std  = float(gap_row[0][1] or 10.0) if gap_row else 10.0
    except Exception:
        avg_gap, gap_std = 45.0, 10.0

    # ── Preferred day-of-week ─────────────────────────────────────────────────
    dow_sql = f"""
    SELECT toDayOfWeek(toDate(date)) AS dow, count() AS cnt
    FROM azure_invoice_report
    WHERE customer_mobile = '{m}'
      AND toDate(date) != toDate('1970-01-01')
    GROUP BY dow
    ORDER BY cnt DESC
    LIMIT 1
    """
    try:
        dow_rows = ch.query(dow_sql).result_rows
        preferred_dow = int(dow_rows[0][0]) if dow_rows else 6
    except Exception:
        preferred_dow = 6

    # ── Branch master enrichment ──────────────────────────────────────────────
    branch_sql = f"""
    SELECT branch_name, district, rbm, bdm, address, store_type
    FROM branch_master
    WHERE code = '{last_branch}'
    LIMIT 1
    """
    try:
        bm = ch.query(branch_sql).result_rows
        branch_name = bm[0][0] if bm else last_branch
        district    = bm[0][1] if bm else ""
        rbm         = bm[0][2] if bm else ""
        bdm         = bm[0][3] if bm else ""
        address     = bm[0][4] if bm else ""
        store_type  = bm[0][5] if bm else ""
    except Exception:
        branch_name = last_branch
        district = rbm = bdm = address = store_type = ""

    # ── Top 3 categories from item_master join ────────────────────────────────
    cat_sql = f"""
    SELECT
        coalesce(im.category, 'OTHER') AS category,
        coalesce(im.brand, '')          AS brand,
        coalesce(im.product, '')        AS product,
        count()                         AS cnt,
        toInt64(sum(s.sold_price))      AS spend
    FROM azure_sales_report s
    LEFT JOIN item_master im ON s.item_code = im.item_code
    WHERE s.customer_mobile = '{m}'
    GROUP BY category, brand, product
    ORDER BY cnt DESC
    LIMIT 3
    """
    try:
        cat_rows = ch.query(cat_sql).result_rows
        top_categories = [
            {"category": cr[0], "brand": cr[1], "product": cr[2],
             "count": int(cr[3]), "spend": int(cr[4])}
            for cr in cat_rows
        ]
    except Exception:
        top_categories = []

    # ── Visit history timeline (last 10 visits) ───────────────────────────────
    hist_sql = f"""
    SELECT toDate(date) AS visit_date, toInt64(sum(invoice_total)) AS spend
    FROM azure_invoice_report
    WHERE customer_mobile = '{m}'
      AND invoice_total > 0
      AND toDate(date) != toDate('1970-01-01')
    GROUP BY visit_date
    ORDER BY visit_date DESC
    LIMIT 10
    """
    try:
        hist_rows = ch.query(hist_sql).result_rows
        visit_history = [
            {"date": h[0].strftime("%d %b %Y"), "spend": int(h[1])}
            for h in hist_rows
        ]
    except Exception:
        visit_history = []

    # ── Build prediction ──────────────────────────────────────────────────────
    window_start  = last_visit + timedelta(days=max(1, int(avg_gap - gap_std)))
    window_end    = last_visit + timedelta(days=int(avg_gap + gap_std))
    peak_date     = _predicted_peak_date(last_visit, avg_gap, preferred_dow)
    days_to_peak  = (peak_date - today).days
    confidence    = _compute_confidence(days_since, avg_gap, gap_std)
    conf_color    = _confidence_color(confidence)
    festival      = _get_festival_in_window(window_start, window_end)

    # Action recommendation
    if confidence == "HIGH":
        action = f"📲 Send WhatsApp reminder today. Customer is due for a visit (peak: {peak_date.strftime('%d %b')})."
    elif confidence == "MEDIUM":
        action = f"📧 Send a value-add offer. Customer is slightly overdue by {days_since - int(avg_gap)} days."
    elif confidence == "AT RISK":
        action = f"🚨 Immediate reactivation needed. Last visit was {days_since} days ago (avg gap: {int(avg_gap)}d)."
    else:
        action = f"⏳ Customer not yet due. Expected to visit around {peak_date.strftime('%d %b %Y')}."

    return {
        "found": True,
        "mobile":           m,
        "mobile_masked":    f"XXXX-{m[-4:]}",
        "last_visit":       last_visit.strftime("%d %b %Y"),
        "days_since":       days_since,
        "visit_count":      visit_count,
        "total_spend":      total_spend,
        "avg_spend":        avg_spend,
        "last_branch":      last_branch,
        "branch_name":      branch_name,
        "district":         district,
        "rbm":              rbm,
        "bdm":              bdm,
        "address":          address,
        "store_type":       store_type,
        "avg_gap":          round(avg_gap, 1),
        "gap_std":          round(gap_std, 1),
        "preferred_dow":    DOW_NAMES.get(preferred_dow, "Saturday"),
        "window_start":     window_start.strftime("%d %b %Y"),
        "window_end":       window_end.strftime("%d %b %Y"),
        "peak_date":        peak_date.strftime("%d %b %Y"),
        "days_to_peak":     days_to_peak,
        "confidence":       confidence,
        "conf_color":       conf_color,
        "festival":         festival or "",
        "action":           action,
        "top_categories":   top_categories,
        "visit_history":    visit_history,
    }


def get_next_visit_excel(predictions: list) -> bytes:
    """
    Generate an Excel file from prediction rows.
    Returns raw bytes of the .xlsx file.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Next Visit Predictions"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1e3a5f")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin     = Side(border_style="thin", color="d1d5db")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Mobile (Masked)", "Last Visit", "Days Since",
        "Visit Count", "Total Spend (₹)", "Avg Spend (₹)",
        "Branch", "District", "RBM",
        "Avg Gap (days)", "Preferred Day",
        "Window Start", "Window End", "Peak Date",
        "Days to Peak", "Confidence", "Festival", "Top Category", "Brand",
    ]
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = 18

    conf_colors = {
        "HIGH":     "d1fae5",
        "MEDIUM":   "fef9c3",
        "AT RISK":  "fee2e2",
        "UPCOMING": "dbeafe",
    }

    for p in predictions:
        row = [
            p["mobile_masked"],
            p["last_visit"],
            p["days_since"],
            p["visit_count"],
            p["total_spend"],
            p["avg_spend"],
            p["branch_name"],
            p["district"],
            p["rbm"],
            p["avg_gap"],
            p["preferred_dow"],
            p["window_start"],
            p["window_end"],
            p["peak_date_short"],
            p["days_to_peak"],
            p["confidence"],
            p["festival"],
            p["category"],
            p["brand"],
        ]
        ws.append(row)
        conf_fill = PatternFill("solid", fgColor=conf_colors.get(p["confidence"], "f9fafb"))
        for cell in ws[ws.max_row]:
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Highlight confidence cell
        conf_cell = ws.cell(row=ws.max_row, column=16)
        conf_cell.fill = conf_fill
        conf_cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
