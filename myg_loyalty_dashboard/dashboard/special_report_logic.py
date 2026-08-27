import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse

def _get_ch():
    from analytics.clickhouse_service import get_ch_client
    return get_ch_client()

def fetch_period_data(ch, start_date, end_date, filters=None):
    """
    Fetch category, product, brand, branch, qty, and value for a specific date range.
    Filters can be single values or comma-separated lists for multi-select.
    """
    if filters is None: filters = {}
    
    def _list(val):
        return [v.strip() for v in val.split(',') if v.strip()] if val else []
    
    def _in_clause(col, val):
        vals = _list(val)
        if not vals: return None
        escaped = "','".join(v.replace("'", "''") for v in vals)
        return f"{col} IN ('{escaped}')"
    
    where_clauses = [f"toDate(s.date) >= toDate('{start_date}')", f"toDate(s.date) <= toDate('{end_date}')"]
    
    from .utils import get_branch_mappings
    code_to_name, name_to_code = get_branch_mappings(ch)

    branch_filter = filters.get('branch', '')
    if branch_filter:
        branch_filter = ','.join([name_to_code.get(b.strip(), b.strip()) for b in branch_filter.split(',')])

    for clause in [
        _in_clause('s.branch',    branch_filter),
        _in_clause('m.category',  filters.get('category', '')),
        _in_clause('m.product',   filters.get('product',  '')),
        _in_clause('m.brand',     filters.get('brand',    '')),
    ]:
        if clause: where_clauses.append(clause)
        
    where_str = " AND ".join(where_clauses)
    
    sql = f"""
        SELECT
            if(isNull(m.category) OR m.category='', 'Unknown', m.category) AS category,
            if(isNull(m.product) OR m.product='', 'Unknown', m.product) AS product,
            if(isNull(m.brand) OR m.brand='', 'Unknown', m.brand)       AS brand,
            if(isNull(s.branch) OR s.branch='', 'Unknown', s.branch)    AS branch,
            SUM(toFloat64(s.qty))                   AS qty,
            SUM(toFloat64(s.sold_price))            AS value
        FROM azure_sales_report s
        LEFT JOIN item_master m ON s.item_code = m.item_code
        WHERE {where_str}
        GROUP BY category, product, brand, branch
        HAVING qty > 0 OR value > 0
    """
    rows = ch.query(sql).result_rows
    keys = ['category', 'product', 'brand', 'branch', 'qty', 'value']
    
    data = []
    for r in rows:
        d = dict(zip(keys, r))
        d['branch'] = code_to_name.get(d['branch'], d['branch'])
        data.append(d)
        
    return data

def build_sheet(ws, data, sheet_title):
    # Setup styling
    THIN = Side(style="thin", color="000000")
    THIN_BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="4F81BD")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")
    
    # Write Headers
    headers = ["Category", "Product", "Brand", "Branch", "QTY", "Value"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = THIN_BDR
        cell.alignment = CENTER
        
    # Sort data for better presentation
    data.sort(key=lambda x: (x['category'], x['product'], x['brand'], x['branch']))
    
    # Write Data very quickly using ws.append
    for r in data:
        ws.append([r['category'], r['product'], r['brand'], r['branch'], r['qty'], r['value']])

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18

def generate_3_period_excel(filters=None):
    wb = Workbook()
    ch = _get_ch()
    
    # Period 1: AMJ 25
    ws_amj25 = wb.active
    ws_amj25.title = "AMJ 25"
    data_amj25 = fetch_period_data(ch, '2025-04-01', '2025-06-30', filters)
    build_sheet(ws_amj25, data_amj25, "AMJ 25")
    
    # Period 2: AMJ 26
    ws_amj26 = wb.create_sheet(title="AMJ 26")
    data_amj26 = fetch_period_data(ch, '2026-04-01', '2026-06-30', filters)
    build_sheet(ws_amj26, data_amj26, "AMJ 26")
    
    # Period 3: JAS 26
    ws_jas26 = wb.create_sheet(title="JAS 26")
    data_jas26 = fetch_period_data(ch, '2026-07-01', '2026-09-30', filters)
    build_sheet(ws_jas26, data_jas26, "JAS 26")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def generate_custom_4_period_excel(filters=None):
    """
    4-sheet report for the user's specific periods:
      Sheet 1: Aug 11-25, 2026
      Sheet 2: Aug 16-Sep 4, 2025
      Sheet 3: Apr 1-May 25, 2025
      Sheet 4: Apr 1-May 25, 2026
    """
    wb = Workbook()
    ch = _get_ch()

    periods = [
        ("Aug 11-25 2026",   '2026-08-11', '2026-08-25'),
        ("Aug 16-Sep4 2025", '2025-08-16', '2025-09-04'),
        ("Apr1-May25 2025",  '2025-04-01', '2025-05-25'),
        ("Apr1-May25 2026",  '2026-04-01', '2026-05-25'),
    ]

    for idx, (title, start, end) in enumerate(periods):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = title
        data = fetch_period_data(ch, start, end, filters)
        build_sheet(ws, data, title)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
