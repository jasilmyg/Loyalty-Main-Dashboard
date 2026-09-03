"""
Export: Mobile-Only Customers (never cross-sold to another retail category)
Fields: Mobile, Name, First Purchase, Last Purchase, Branch, Total Invoices, Total Spend
"""
import os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'myg_loyalty_dashboard.settings')
sys.path.insert(0, '.')
django.setup()

from analytics.clickhouse_service import get_ch_client
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

ch = get_ch_client()

EXCLUDE = """'STATIONERY ITEMS','SCHEME','GDOT CARE','D SPARE','OSG WARRANTY',
             'SERVICE','TOTAL SECURITY','LG AMC','SERVICE CHARGES',
             'DEMO','DEMO LAPTOP','DEMO ACCESSORIES','MYG DOMO','MYG VERSE',
             'DIY','CONTRACT WORK','CEGI','RIG','PROTECT MAX','CARE PLUS',
             'MOBILE ANTIVIRUS','HA ACCESSORIES','CCTV','MONITOR'"""

print("Step 1: Fetching mobile-only customers... (3–5 min)")

rows = ch.query(f"""
    WITH mobile_buyers AS (
        SELECT
            ai.customer_mobile               AS mobile,
            any(ai.customer_type)            AS cust_type,
            min(toDate(ai.date))             AS first_date,
            max(toDate(ai.date))             AS last_date,
            countDistinct(ai.invoice_no)     AS total_invoices,
            sum(toFloat64(sr.sold_price))    AS total_spend,
            any(ai.branch)                   AS branch_code,
            any(ai.rbm)                      AS rbm,
            any(ai.bdm)                      AS bdm
        FROM azure_invoice_report ai
        INNER JOIN azure_sales_report sr ON ai.invoice_no = sr.invoice_no
        INNER JOIN item_master m         ON sr.item_code = m.item_code
        WHERE m.product = 'MOBILE'
          AND length(trim(ai.customer_mobile)) >= 10
          AND toDate(ai.date) BETWEEN '2021-01-01' AND '2026-08-29'
          AND toDate(ai.date) != '1970-01-01'
        GROUP BY ai.customer_mobile
    ),
    cross_shoppers AS (
        SELECT DISTINCT ai2.customer_mobile AS mobile
        FROM azure_invoice_report ai2
        INNER JOIN azure_sales_report sr2 ON ai2.invoice_no = sr2.invoice_no
        INNER JOIN item_master m2         ON sr2.item_code = m2.item_code
        INNER JOIN mobile_buyers mb       ON ai2.customer_mobile = mb.mobile
        WHERE m2.product NOT IN ({EXCLUDE})
          AND m2.product != 'MOBILE'
          AND toDate(ai2.date) > mb.first_date
          AND toDate(ai2.date) != '1970-01-01'
          AND length(trim(ai2.customer_mobile)) >= 10
    )
    SELECT
        mb.mobile,
        mb.cust_type,
        mb.first_date,
        mb.last_date,
        mb.total_invoices,
        mb.total_spend,
        mb.branch_code,
        b.branch_name,
        mb.rbm,
        mb.bdm
    FROM mobile_buyers mb
    LEFT JOIN branch_master b ON mb.branch_code = b.code
    WHERE mb.mobile NOT IN (SELECT mobile FROM cross_shoppers)
    ORDER BY mb.total_spend DESC
""").result_rows

print(f"  → Found {len(rows):,} mobile-only customers")
print("Step 2: Building DataFrame...")

df = pd.DataFrame(rows, columns=[
    'Mobile', 'Customer Type', 'First Purchase Date', 'Last Purchase Date',
    'Total Invoices', 'Total Mobile Spend (₹)',
    'Branch Code', 'Branch Name', 'RBM', 'BDM'
])

df['First Purchase Date'] = pd.to_datetime(df['First Purchase Date']).dt.date
df['Last Purchase Date']  = pd.to_datetime(df['Last Purchase Date']).dt.date
df['Total Mobile Spend (₹)'] = df['Total Mobile Spend (₹)'].round(0).astype(int)
df['Days Since Last Purchase'] = (
    pd.to_datetime('2026-08-29') - pd.to_datetime(df['Last Purchase Date'])
).dt.days

print(f"Step 3a: Saving full data as CSV... ({len(df):,} rows)")

CSV_FILE = r'C:\Users\jasil_myg\Desktop\Mobile_Only_Customers_FULL.csv'
df_export = df.drop(columns=['Branch Code'])
df_export.to_csv(CSV_FILE, index=False, encoding='utf-8-sig')
print(f"  ✅ CSV saved → {CSV_FILE}")

# ── Summary Excel ─────────────────────────────────────────────────────────────
print("Step 3b: Building Summary Excel...")

OUTFILE = r'C:\Users\jasil_myg\Desktop\Mobile_Only_Customers_SUMMARY.xlsx'
wb = Workbook()

# ── Helper ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1E3A5F')
HDR_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
ALT_FILL  = PatternFill('solid', fgColor='EEF4FB')
BODY_FONT = Font(name='Calibri', size=9)
THIN      = Side(style='thin', color='CCCCCC')
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER    = Alignment(horizontal='center', vertical='center')
RIGHT_A   = Alignment(horizontal='right',  vertical='center')
LEFT_A    = Alignment(horizontal='left',   vertical='center')

def write_sheet(ws, title, headers_widths, data_rows, fmt_map=None):
    ws.title = title
    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(headers_widths))}1')
    ws['A1'] = f'  myG — Mobile-Only Customers: {title}'
    ws['A1'].font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    ws['A1'].fill = PatternFill('solid', fgColor='1E3A5F')
    ws['A1'].alignment = LEFT_A
    ws.row_dimensions[1].height = 24
    # Headers
    ws.row_dimensions[2].height = 26
    for ci, (h, w) in enumerate(headers_widths, 1):
        c = ws.cell(2, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    # Data
    for ri, row in enumerate(data_rows, 3):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font = BODY_FONT; c.border = BORDER
            if fill: c.fill = fill
            if fmt_map and ci in fmt_map:
                c.number_format = fmt_map[ci]
                c.alignment = RIGHT_A
            else:
                c.alignment = LEFT_A
        ws.row_dimensions[ri].height = 15
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers_widths))}{2+len(data_rows)}'

# ── Sheet 1: Overview ────────────────────────────────────────────────────────
ws0 = wb.active
ws0.title = 'Overview'
ws0['A1'] = 'myG — Mobile-Only Customers Report'
ws0['A1'].font = Font(name='Calibri', bold=True, size=14, color='1E3A5F')
ws0.row_dimensions[1].height = 28

stats = [
    ('Total Mobile-Only Customers',       f"{len(df):,}"),
    ('Total Mobile Spend (₹)',            f"₹{df['Total Mobile Spend (₹)'].sum():,.0f}"),
    ('Avg Spend per Customer (₹)',        f"₹{df['Total Mobile Spend (₹)'].mean():,.0f}"),
    ('Median Spend per Customer (₹)',     f"₹{df['Total Mobile Spend (₹)'].median():,.0f}"),
    ('Avg Total Invoices per Customer',   f"{df['Total Invoices'].mean():.1f}"),
    ('Avg Days Since Last Purchase',      f"{df['Days Since Last Purchase'].mean():.0f} days"),
    ('Customers inactive >365 days',      f"{(df['Days Since Last Purchase']>365).sum():,}  ({(df['Days Since Last Purchase']>365).mean()*100:.1f}%)"),
    ('Customers inactive >180 days',      f"{(df['Days Since Last Purchase']>180).sum():,}  ({(df['Days Since Last Purchase']>180).mean()*100:.1f}%)"),
    ('Customers inactive <90 days',       f"{(df['Days Since Last Purchase']<90).sum():,}  ({(df['Days Since Last Purchase']<90).mean()*100:.1f}%)"),
    ('Report Generated',                  '29-Aug-2026'),
    ('Full Data',                         'Mobile_Only_Customers_FULL.csv (on Desktop)'),
]
for i, (k, v) in enumerate(stats, 2):
    ws0.cell(i, 1, k).font = Font(name='Calibri', bold=True, size=10, color='1E3A5F')
    ws0.cell(i, 2, v).font = Font(name='Calibri', size=10)
    ws0.column_dimensions['A'].width = 40
    ws0.column_dimensions['B'].width = 45

# ── Sheet 2: Branch Summary ──────────────────────────────────────────────────
branch_grp = df.groupby('Branch Name').agg(
    Customers=('Mobile','count'),
    Total_Spend=('Total Mobile Spend (₹)','sum'),
    Avg_Spend=('Total Mobile Spend (₹)','mean'),
    Avg_Days_Inactive=('Days Since Last Purchase','mean'),
    Avg_Invoices=('Total Invoices','mean'),
).reset_index().sort_values('Customers', ascending=False)
branch_grp['% of Total'] = branch_grp['Customers'] / len(df) * 100

ws1 = wb.create_sheet()
write_sheet(ws1, 'By Branch', [
    ('Branch Name',22),('Customers',12),('% of Total',11),
    ('Total Spend (₹)',18),('Avg Spend/Cust (₹)',18),
    ('Avg Days Inactive',16),('Avg Invoices',12)
], [
    (r['Branch Name'], int(r['Customers']), round(r['% of Total'],2),
     round(r['Total_Spend'],0), round(r['Avg_Spend'],0),
     round(r['Avg_Days_Inactive'],0), round(r['Avg_Invoices'],1))
    for _, r in branch_grp.iterrows()
], fmt_map={3:'0.00%', 4:'#,##0', 5:'#,##0'})

# ── Sheet 3: RBM Summary ─────────────────────────────────────────────────────
rbm_grp = df.groupby('RBM').agg(
    Customers=('Mobile','count'),
    Total_Spend=('Total Mobile Spend (₹)','sum'),
    Avg_Spend=('Total Mobile Spend (₹)','mean'),
    Avg_Days_Inactive=('Days Since Last Purchase','mean'),
).reset_index().sort_values('Customers', ascending=False)
rbm_grp['% of Total'] = rbm_grp['Customers'] / len(df) * 100

ws2 = wb.create_sheet()
write_sheet(ws2, 'By RBM', [
    ('RBM',20),('Customers',12),('% of Total',11),
    ('Total Spend (₹)',18),('Avg Spend/Cust (₹)',18),('Avg Days Inactive',16)
], [
    (r['RBM'], int(r['Customers']), round(r['% of Total'],2),
     round(r['Total_Spend'],0), round(r['Avg_Spend'],0),
     round(r['Avg_Days_Inactive'],0))
    for _, r in rbm_grp.iterrows()
], fmt_map={4:'#,##0', 5:'#,##0'})

# ── Sheet 4: Inactivity Buckets ──────────────────────────────────────────────
ws3 = wb.create_sheet('Inactivity Buckets')
ws3['A1'] = 'myG — Mobile-Only Customers: Inactivity Segments'
ws3['A1'].font = Font(name='Calibri', bold=True, size=12, color='1E3A5F')
ws3.row_dimensions[1].height = 24
buckets = [
    ('< 30 days',   df['Days Since Last Purchase'].lt(30).sum()),
    ('30–90 days',  df['Days Since Last Purchase'].between(30,90).sum()),
    ('91–180 days', df['Days Since Last Purchase'].between(91,180).sum()),
    ('181–365 days',df['Days Since Last Purchase'].between(181,365).sum()),
    ('1–2 years',   df['Days Since Last Purchase'].between(366,730).sum()),
    ('2–3 years',   df['Days Since Last Purchase'].between(731,1095).sum()),
    ('> 3 years',   df['Days Since Last Purchase'].gt(1095).sum()),
]
headers = [('Inactivity Bucket',22),('Customers',14),('% of Total',12),('Action Priority',20)]
for ci, (h, w) in enumerate(headers, 1):
    c = ws3.cell(2, ci, h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[2].height = 22
priority_map = {
    '< 30 days':'🟢 Recently active — nurture',
    '30–90 days':'🟢 Warm — send cross-sell offer now',
    '91–180 days':'🟡 At risk — send reactivation',
    '181–365 days':'🟡 Cold — discount offer needed',
    '1–2 years':'🔴 Lapsed — strong incentive needed',
    '2–3 years':'🔴 Highly lapsed — last-chance campaign',
    '> 3 years':'⚫ Likely churned permanently',
}
total = len(df)
for ri, (label, count) in enumerate(buckets, 3):
    fill = ALT_FILL if ri % 2 == 0 else None
    for ci, val in enumerate([label, count, round(count/total*100,1), priority_map[label]], 1):
        c = ws3.cell(ri, ci, val)
        c.font = BODY_FONT; c.border = BORDER
        if fill: c.fill = fill
        c.alignment = LEFT_A
    ws3.row_dimensions[ri].height = 16

wb.save(OUTFILE)
print(f"\n✅ Summary Excel saved → {OUTFILE}")
print(f"✅ Full CSV saved      → {CSV_FILE}")
print(f"\n   CSV rows     : {len(df):,}")
print(f"   Excel sheets : Overview | By Branch | By RBM | Inactivity Buckets")
