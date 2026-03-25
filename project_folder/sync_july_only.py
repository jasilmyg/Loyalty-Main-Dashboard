import sys, os, io, sqlite3, time
sys.stdout.reconfigure(line_buffering=True)

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

SERVICE_ACCOUNT_FILE = 'service_account.json'
FOLDER_ID = '1YDtp98SA2HLTbFjkTzOZy0R28fJmz2kL'
DB_NAME = 'combined_data.db'
TABLE_NAME = 'sales_data'

def get_service():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=['https://www.googleapis.com/auth/drive.readonly']
    )
    return build('drive', 'v3', credentials=creds, static_discovery=False)

def list_july_files(service):
    files = []
    page_token = None
    while True:
        resp = service.files().list(
            q=f"'{FOLDER_ID}' in parents and trashed=false and name contains 'JULY 2025'",
            spaces='drive',
            fields='nextPageToken, files(id, name)',
            pageToken=page_token
        ).execute()
        for f in resp.get('files', []):
            if f['name'].endswith('.xlsx'):
                files.append(f)
        page_token = resp.get('nextPageToken')
        if not page_token:
            break
    return files

def download_file(service, file_id):
    req = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf

def read_excel_rows(buf):
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = None
    data_rows = []
    for row in rows_iter:
        if headers is None:
            headers = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(row)]
            continue
        data_rows.append(row)
    wb.close()
    return headers, data_rows

def insert_rows(conn, headers, rows, source_file):
    all_headers = headers + ['source_file']
    placeholders = ', '.join(['?'] * len(all_headers))
    col_names = ', '.join(f'[{h}]' for h in all_headers)
    
    # Ensure source_file column exists
    try:
        conn.execute(f'ALTER TABLE [{TABLE_NAME}] ADD COLUMN [source_file] TEXT')
        conn.commit()
    except:
        pass
    
    # Ensure all columns from this file exist in the table
    for h in headers:
        try:
            conn.execute(f'ALTER TABLE [{TABLE_NAME}] ADD COLUMN [{h}] TEXT')
            conn.commit()
        except:
            pass
            
    # Check if this file was already imported previously to avoid duplicates if rerun
    count = conn.execute(f"SELECT COUNT(*) FROM [{TABLE_NAME}] WHERE [source_file] = ?", (source_file,)).fetchone()[0]
    if count > 0:
        print(f"    Warning: Deleting {count} existing rows for {source_file} to prevent duplicates.")
        conn.execute(f"DELETE FROM [{TABLE_NAME}] WHERE [source_file] = ?", (source_file,))
        conn.commit()
    
    batch = []
    for row in rows:
        values = list(row) + [source_file]
        while len(values) < len(all_headers):
            values.append(None)
        values = values[:len(all_headers)]
        batch.append(tuple(str(v) if v is not None else None for v in values))
    
    if batch:
        conn.executemany(
            f'INSERT INTO [{TABLE_NAME}] ({col_names}) VALUES ({placeholders})',
            batch
        )
        conn.commit()
    return len(batch)

def main():
    print("Loading credentials...", flush=True)
    service = get_service()
    
    print("Fetching JULY 2025 file list...", flush=True)
    files = list_july_files(service)
    print(f"Found {len(files)} matching Excel files.", flush=True)
    
    if not files:
        return
    
    # DO NOT DROP TABLE. WE APPEND INSTEAD.
    conn = sqlite3.connect(DB_NAME)
    
    total = len(files)
    total_rows = 0
    
    for idx, f in enumerate(files, 1):
        name = f['name']
        fid = f['id']
        print(f"\n[{idx}/{total}] {name}", flush=True)
        
        try:
            t0 = time.time()
            buf = download_file(service, fid)
            dl_t = time.time() - t0
            print(f"  Downloaded ({dl_t:.1f}s). Reading...", flush=True)
            
            t1 = time.time()
            headers, rows = read_excel_rows(buf)
            read_t = time.time() - t1
            
            if not rows:
                print(f"  Empty file, skipping.", flush=True)
                continue
            
            t2 = time.time()
            n = insert_rows(conn, headers, rows, name)
            write_t = time.time() - t2
            total_rows += n
            print(f"  {n} rows. [DL:{dl_t:.1f}s Read:{read_t:.1f}s Write:{write_t:.1f}s] Total Added: {total_rows}", flush=True)
            
        except Exception as e:
            print(f"  ERROR: {e}", flush=True)
    
    conn.close()
    print(f"\nDone! Appended {total_rows} total rows.", flush=True)

if __name__ == '__main__':
    main()
